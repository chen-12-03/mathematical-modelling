import math
import numpy as np
import random
import itertools
import functools

def calculate_smoke_valid_duration(drone_speed, theta_deg, t_drop, t_burst, fyx_pos_init, step=0.01):
    # 常量定义
    fyx_point=fyx_pos_init
    fake_point=(0, 0, 0)
    g=9.8
    smoke_center_vz=-3
    smoke_valid_time=20

    # --- FYX飞行部分---
    vx = drone_speed * np.cos(np.radians(theta_deg))
    vy = drone_speed * np.sin(np.radians(theta_deg))
    vz = 0

    delta_x = vx * t_drop
    delta_y = vy * t_drop
    delta_z = vz * t_drop

    new_x_drone = fyx_point[0] + delta_x
    new_y_drone = fyx_point[1] + delta_y
    new_z_drone = fyx_point[2] + delta_z

    # --- 烟雾弹下落部分 ---
    smoke_bomb_initial_vx = vx
    smoke_bomb_initial_vy = vy
    smoke_bomb_initial_vz = vz

    delta_vz_gravity = -g * t_burst

    # smoke_bomb在z方向的最终速度
    smoke_bomb_vz_final = smoke_bomb_initial_vz + delta_vz_gravity

    # 计算smoke_bomb下落smoke_bomb_fall_time秒后的位置
    smoke_bomb_initial_x = new_x_drone
    smoke_bomb_initial_y = new_y_drone
    smoke_bomb_initial_z = new_z_drone

    delta_x_smoke_bomb = smoke_bomb_initial_vx * t_burst
    delta_y_smoke_bomb = smoke_bomb_initial_vy * t_burst
    delta_z_smoke_bomb = smoke_bomb_initial_vz * t_burst + 0.5 * (-g) * t_burst**2

    smoke_bomb_final_x = smoke_bomb_initial_x + delta_x_smoke_bomb
    smoke_bomb_final_y = smoke_bomb_initial_y + delta_y_smoke_bomb
    smoke_bomb_final_z = smoke_bomb_initial_z + delta_z_smoke_bomb

    # --- smoke_center运动部分 ---
    smoke_center_vx = 0
    smoke_center_vy = 0

    smoke_center_start_time_total = t_drop + t_burst

    def calculate_smoke_center_position(total_time_t):
        if total_time_t < smoke_center_start_time_total:
            return (smoke_bomb_final_x, smoke_bomb_final_y, smoke_bomb_final_z)
        else:
            smoke_center_motion_duration = total_time_t - smoke_center_start_time_total
            sc_x = smoke_bomb_final_x + smoke_center_vx * smoke_center_motion_duration
            sc_y = smoke_bomb_final_y + smoke_center_vy * smoke_center_motion_duration
            sc_z = smoke_bomb_final_z + smoke_center_vz * smoke_center_motion_duration
            return (sc_x, sc_y, sc_z)

    # --- M1运动部分 ---
    M1_initial_point = (20000, 0, 2000)
    M1_speed = 300

    M1_direction_vector = (
        fake_point[0] - M1_initial_point[0],
        fake_point[1] - M1_initial_point[1],
        fake_point[2] - M1_initial_point[2]
    )
    M1_distance = math.sqrt(
        M1_direction_vector[0]**2 + M1_direction_vector[1]**2 + M1_direction_vector[2]**2
    )
    M1_unit_direction_vector = (
        M1_direction_vector[0] / M1_distance,
        M1_direction_vector[1] / M1_distance,
        M1_direction_vector[2] / M1_distance
    )

    M1_vx = M1_speed * M1_unit_direction_vector[0]
    M1_vy = M1_speed * M1_unit_direction_vector[1]
    M1_vz = M1_speed * M1_unit_direction_vector[2]

    def calculate_M1_position(total_time_t):
        M1_delta_x = M1_vx * total_time_t
        M1_delta_y = M1_vy * total_time_t
        M1_delta_z = M1_vz * total_time_t
        M1_final_x = M1_initial_point[0] + M1_delta_x
        M1_final_y = M1_initial_point[1] + M1_delta_y
        M1_final_z = M1_initial_point[2] + M1_delta_z
        return (M1_final_x, M1_final_y, M1_final_z)

    def point_to_segment_distance(segment_start, segment_end, point):
        def distance_between_points(p1, p2):
            return math.sqrt((p2[0] - p1[0])**2 + (p2[1] - p1[1])**2 + (p2[2] - p1[2])**2)
        def dot_product(v1, v2):
            return v1[0] * v2[0] + v1[1] * v2[1] + v1[2] * v2[2]
        def vector_subtract(p1, p2):
            return (p1[0] - p2[0], p1[1] - p2[1], p1[2] - p2[2])

        AB = vector_subtract(segment_end, segment_start)
        AP = vector_subtract(point, segment_start)
        AB_squared = dot_product(AB, AB)
        if AB_squared == 0:
            return distance_between_points(point, segment_start)
        t = dot_product(AP, AB) / AB_squared
        if t < 0:
            return distance_between_points(point, segment_start)
        elif t > 1:
            return distance_between_points(point, segment_end)
        else:
            projection_point = (
                segment_start[0] + t * AB[0],
                segment_start[1] + t * AB[1],
                segment_start[2] + t * AB[2]
            )
            return distance_between_points(point, projection_point)

    real_point_1 = (7, 200, 0)
    real_point_2 = (-7, 200, 10)

    smoke_valid_duration = 0

    start_loop_time = smoke_center_start_time_total
    end_loop_time = smoke_center_start_time_total + smoke_valid_time + step
    
    t_valid_begin = 0
    t_valid_end = 0
    smoke_valid_duration = 0
    for t in np.arange(start_loop_time, end_loop_time, step):
        smoke_center_pos = calculate_smoke_center_position(t)
        M1_pos = calculate_M1_position(t)
        dist_1 = point_to_segment_distance(real_point_1, M1_pos, smoke_center_pos)
        dist_2 = point_to_segment_distance(real_point_2, M1_pos, smoke_center_pos)
        if dist_1 < 10 and dist_2 < 10:
            smoke_valid_duration += step
            if t_valid_begin == 0:
                t_valid_begin = t
        if dist_1 >=10 and dist_2 >= 10 and t_valid_begin != 0:
            t_valid_end = t
            break

    t_valid_interval = (t_valid_begin, t_valid_end)
    drop_pos = (new_x_drone, new_y_drone, new_z_drone)
    burst_pos = (smoke_bomb_final_x, smoke_bomb_final_y, smoke_bomb_final_z)
    smoke_inf = (drop_pos, burst_pos, smoke_valid_duration)

    return smoke_valid_duration, t_valid_interval, smoke_inf

# 无人机初始位置定义
fy1_point = (17800, 0, 1800)
fy2_point = (12000, 1400, 1400)
fy3_point = (6000, -3000, 700)
fy4_point = (11000, 2000, 1800)
fy5_point = (13000, -2000, 1300)
fyx_points = [fy1_point, fy2_point, fy3_point, fy4_point, fy5_point]

def union_interval_length(intervals):
    # intervals: [(start1, end1), (start2, end2), ...]
    # 先按起点排序
    intervals = sorted(intervals, key=lambda x: x[0])
    merged = []
    for interval in intervals:
        if not merged or merged[-1][1] < interval[0]:
            merged.append(list(interval))
        else:
            merged[-1][1] = max(merged[-1][1], interval[1])
    # 计算总长度
    total = sum(end - start for start, end in merged)
    return total

def set_all_seeds(seed):
    random.seed(seed)
    np.random.seed(seed)

# SEED_VALUE = 6666
SEED_VALUE = 30

set_all_seeds(SEED_VALUE)

from sko.PSO import PSO
def demo_func(params, fyx_point, lb, ub):
    params = np.clip(params, lb, ub)
    total_duration, *_ = calculate_smoke_valid_duration(*params, fyx_pos_init=fyx_point)
    return -total_duration

# ai生成用于限制PSO SA求解取值范围
# 包装器函数：边界检查+惩罚
def objective_function_wrapper(params, fyx_point, lb, ub):
    params = np.array(params)
    lb = np.array(lb)
    ub = np.array(ub)
    # 边界检查
    if np.any(params < lb) or np.any(params > ub):
        return 1e10  # 越界惩罚
    total_duration, *_ = calculate_smoke_valid_duration(*params, fyx_pos_init=fyx_point)
    return -total_duration

# PSO部分（可选用包装器，便于统一）
print("--- Running PSO ---")
pso_1 = PSO(
    func=lambda params: objective_function_wrapper(params, fyx_points[0], [70, 0, 0, 0], [140, 180, 5, 5]),
    dim=4, pop=75, max_iter=70,
    lb=[70, 0, 0, 0], ub=[140, 180, 5, 5],
    w=0.8, c1=0.5, c2=0.5
)
pso_1.run()

pso_2 = PSO(
    func=lambda params: objective_function_wrapper(params, fyx_points[1], [70, 225, 0, 0], [140, 235, 40, 40]),
    dim=4, pop=200, max_iter=70,
    lb=[70, 180, 0, 0], ub=[140, 360, 40, 40],
    w=0.8, c1=0.5, c2=0.5
)
pso_2.run()

pso_3 = PSO(
    func=lambda params: objective_function_wrapper(params, fyx_points[2], [70, 0, 0, 0], [140, 180, 50, 50]),
    dim=4, pop=200, max_iter=70,
    lb=[70, 0, 0, 0], ub=[140, 180, 50, 50],
    w=0.8, c1=0.5, c2=0.5
)
pso_3.run()

# SA部分（必须用包装器！）
print("\n--- Running SA with Penalty Function ---")
from sko.SA import SA

lb1 = np.array([70, 0, 0, 0])
ub1 = np.array([140, 180, 5, 5])
lb2 = np.array([70, 225, 0, 0])
ub2 = np.array([140, 235, 30, 30])
lb3 = np.array([70, 85, 0, 0])
ub3 = np.array([140, 105, 40, 40])

func1 = functools.partial(objective_function_wrapper, fyx_point=fyx_points[0], lb=lb1, ub=ub1)
func2 = functools.partial(objective_function_wrapper, fyx_point=fyx_points[1], lb=lb2, ub=ub2)
func3 = functools.partial(objective_function_wrapper, fyx_point=fyx_points[2], lb=lb3, ub=ub3)

sa_1 = SA(func=func1, x0=pso_1.gbest_x, T_max=10, T_min=1e-9, L=100, max_stay_counter=150, lb=lb1, ub=ub1)
sa_2 = SA(func=func2, x0=pso_2.gbest_x, T_max=10, T_min=1e-9, L=100, max_stay_counter=150, lb=lb2, ub=ub2)
sa_3 = SA(func=func3, x0=pso_3.gbest_x, T_max=10, T_min=1e-9, L=100, max_stay_counter=150, lb=lb3, ub=ub3)

best_x1, best_y1 = sa_1.run()
formatted_sa_x1 = [round(float(x), 3) for x in best_x1]
formatted_sa_y1 = round(float(-best_y1), 2)
print('SA1 最佳参数:', formatted_sa_x1, '最佳值:', formatted_sa_y1)

best_x2, best_y2 = sa_2.run()
formatted_sa_x2 = [round(float(x), 3) for x in best_x2]
formatted_sa_y2 = round(float(-best_y2), 2)
print('SA2 最佳参数:', formatted_sa_x2, '最佳值:', formatted_sa_y2)

best_x3, best_y3 = sa_3.run()
formatted_sa_x3 = [round(float(x), 3) for x in best_x3]
formatted_sa_y3 = round(float(-best_y3), 2)
print('SA3 最佳参数:', formatted_sa_x3, '最佳值:', formatted_sa_y3)

# 计算每组SA最优参数对应的遮蔽区间和时长
t_duration_1, t_valid_interval_1, smoke_inf_1 = calculate_smoke_valid_duration(*best_x1, fyx_points[0])
t_span_1 = [t_valid_interval_1[0], t_valid_interval_1[1]]

t_duration_2, t_valid_interval_2, smoke_inf_2 = calculate_smoke_valid_duration(*best_x2, fyx_points[1])
t_span_2 = [t_valid_interval_2[0], t_valid_interval_2[1]]

t_duration_3, t_valid_interval_3, smoke_inf_3 = calculate_smoke_valid_duration(*best_x3, fyx_points[2])
t_span_3 = [t_valid_interval_3[0], t_valid_interval_3[1]]

# 重要参数计算
intervals = [t_span_1, t_span_2, t_span_3]
t_spans = [t_duration_1, t_duration_2, t_duration_3]
print(f"t_duration_1 = {t_span_1[1] - t_span_1[0]:.2f}, t_span_1 = [{t_span_1[0]:.2f}, {t_span_1[1]:.2f}]")
print(f"t_duration_2 = {t_span_2[1] - t_span_2[0]:.2f}, t_span_2 = [{t_span_2[0]:.2f}, {t_span_2[1]:.2f}]")
print(f"t_duration_3 = {t_span_3[1] - t_span_3[0]:.2f}, t_span_3 = [{t_span_3[0]:.2f}, {t_span_3[1]:.2f}]")
total_duration = union_interval_length(intervals)
print(f"总遮蔽时长: {total_duration:.2f} s")

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

excel_file_path = 'result2.xlsx'

def fmt(num, x):
    # 保留三位小数，如果不是数字则原样返回
    try:
        return round(float(num), x)
    except Exception:
        return num

data_to_fill = {
    '无人机运动方向': [fmt(best_x1[1], 3), fmt(best_x2[1], 3), fmt(best_x3[1], 3)],
    '无人机运动速度 (m/s)': [fmt(best_x1[0], 3), fmt(best_x2[0], 3), fmt(best_x3[0], 3)],
    '烟幕干扰弹投放点的x坐标 (m)': [fmt(smoke_inf_1[0][0], 3), fmt(smoke_inf_2[0][0], 3), fmt(smoke_inf_3[0][0], 3)],
    '烟幕干扰弹投放点的y坐标 (m)': [fmt(smoke_inf_1[0][1], 3), fmt(smoke_inf_2[0][1], 3), fmt(smoke_inf_3[0][1], 3)],
    '烟幕干扰弹投放点的z坐标 (m)': [fmt(smoke_inf_1[0][2], 3), fmt(smoke_inf_2[0][2], 3), fmt(smoke_inf_3[0][2], 3)],
    '烟幕干扰弹起爆点的x坐标 (m)': [fmt(smoke_inf_1[1][0], 3), fmt(smoke_inf_2[1][0], 3), fmt(smoke_inf_3[1][0], 3)],
    '烟幕干扰弹起爆点的y坐标 (m)': [fmt(smoke_inf_1[1][1], 3), fmt(smoke_inf_2[1][1], 3), fmt(smoke_inf_3[1][1], 3)],
    '烟幕干扰弹起爆点的z坐标 (m)': [fmt(smoke_inf_1[1][2], 3), fmt(smoke_inf_2[1][2], 3), fmt(smoke_inf_3[1][2], 3)],
    '有效干扰时长 (s)': [fmt(t_spans[0], 2), fmt(t_spans[1], 2), fmt(t_spans[2], 2)],
}

# 在保存Excel前加上用户确认
user_input = input("是否保存结果到Excel？输入Y或y保存，其他任意键不保存：")
if user_input.strip().lower() == 'y':
    workbook = load_workbook(excel_file_path)
    sheet = workbook.active

    # 获取列标题到列索引的映射
    column_map = {}
    for col_idx, cell in enumerate(sheet[1], 1):
        column_map[cell.value] = get_column_letter(col_idx)

    # 确定数据填充的起始行（第一行为标题行）
    start_row_index = 2 # Excel 行从 1 开始，标题占一行，数据从第 2 行开始

    # 准备要填充的数据，确保所有列表长度一致
    max_data_rows = 0
    for col_data in data_to_fill.values():
        max_data_rows = max(max_data_rows, len(col_data))

    for col_name in data_to_fill:
        while len(data_to_fill[col_name]) < max_data_rows:
            data_to_fill[col_name].append(None)

    # 填充数据
    for col_header, col_data in data_to_fill.items():
        if col_header in column_map:
            column_letter = column_map[col_header]
            for i, value in enumerate(col_data):
                target_row = start_row_index + i
                cell = sheet[f"{column_letter}{target_row}"]
                cell.value = value
        else:
            print(f"警告：列 '{col_header}' 在 Excel 文件中未找到，数据将不会填充。")
            
    # 保存修改后的 Excel 文件
    workbook.save(excel_file_path)
    print(f"成功使用新数据更新了 '{excel_file_path}'。")
else:
    print("未保存结果到Excel。")