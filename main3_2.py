import math
import numpy as np
import random
from sko.GA import GA
from sko.SA import SA # 建议也导入SA，用于混合策略

# ===================================================================
# 你的物理模型和辅助函数 (这部分保持不变)
# ===================================================================
def calculate_smoke_valid_duration(drone_speed, theta_deg, t_drop, t_burst, fyx_pos_init, step=0.01, t_bias=0):
    # ... (你的完整物理模型代码) ...
    # (为了简洁，此处省略函数体)
    # 常量定义
    fyx_point=fyx_pos_init
    fake_point=(0, 0, 0)
    g=9.8
    smoke_center_vz=-3
    smoke_valid_time=20

    # --- FY1飞行部分---
    vx = drone_speed * np.cos(np.radians(theta_deg))
    vy = drone_speed * np.sin(np.radians(theta_deg))
    vz = 0

    delta_x = vx * t_drop
    delta_y = vy * t_drop
    delta_z = vz * t_drop

    new_x_drone = fyx_point[0] + delta_x
    new_y_drone = fyx_point[1] + delta_y
    new_z_drone = fyx_point[2] + delta_z

    # --- 烟雾弹下落部分 ---
    smoke_bomb_initial_vx = vx
    smoke_bomb_initial_vy = vy
    smoke_bomb_initial_vz = vz

    delta_vz_gravity = -g * t_burst

    # smoke_bomb在z方向的最终速度
    smoke_bomb_vz_final = smoke_bomb_initial_vz + delta_vz_gravity

    # 计算smoke_bomb下落smoke_bomb_fall_time秒后的位置
    smoke_bomb_initial_x = new_x_drone
    smoke_bomb_initial_y = new_y_drone
    smoke_bomb_initial_z = new_z_drone

    delta_x_smoke_bomb = smoke_bomb_initial_vx * t_burst
    delta_y_smoke_bomb = smoke_bomb_initial_vy * t_burst
    delta_z_smoke_bomb = smoke_bomb_initial_vz * t_burst + 0.5 * (-g) * t_burst**2

    smoke_bomb_final_x = smoke_bomb_initial_x + delta_x_smoke_bomb
    smoke_bomb_final_y = smoke_bomb_initial_y + delta_y_smoke_bomb
    smoke_bomb_final_z = smoke_bomb_initial_z + delta_z_smoke_bomb

    # --- smoke_center运动部分 ---
    smoke_center_vx = 0
    smoke_center_vy = 0

    smoke_center_start_time_total = t_drop + t_burst

    def calculate_smoke_center_position(total_time_t):
        if total_time_t < smoke_center_start_time_total:
            return (smoke_bomb_final_x, smoke_bomb_final_y, smoke_bomb_final_z)
        else:
            smoke_center_motion_duration = total_time_t - smoke_center_start_time_total
            sc_x = smoke_bomb_final_x + smoke_center_vx * smoke_center_motion_duration
            sc_y = smoke_bomb_final_y + smoke_center_vy * smoke_center_motion_duration
            sc_z = smoke_bomb_final_z + smoke_center_vz * smoke_center_motion_duration
            return (sc_x, sc_y, sc_z)

    # --- M1运动部分 ---
    M1_initial_point = (20000, 0, 2000)
    M1_speed = 300

    M1_direction_vector = (
        fake_point[0] - M1_initial_point[0],
        fake_point[1] - M1_initial_point[1],
        fake_point[2] - M1_initial_point[2]
    )
    M1_distance = math.sqrt(
        M1_direction_vector[0]**2 + M1_direction_vector[1]**2 + M1_direction_vector[2]**2
    )
    M1_unit_direction_vector = (
        M1_direction_vector[0] / M1_distance,
        M1_direction_vector[1] / M1_distance,
        M1_direction_vector[2] / M1_distance
    )

    M1_vx = M1_speed * M1_unit_direction_vector[0]
    M1_vy = M1_speed * M1_unit_direction_vector[1]
    M1_vz = M1_speed * M1_unit_direction_vector[2]

    def calculate_M1_position(total_time_t):
        t_m1 = total_time_t + t_bias
        M1_delta_x = M1_vx * t_m1
        M1_delta_y = M1_vy * t_m1
        M1_delta_z = M1_vz * t_m1
        M1_final_x = M1_initial_point[0] + M1_delta_x
        M1_final_y = M1_initial_point[1] + M1_delta_y
        M1_final_z = M1_initial_point[2] + M1_delta_z
        return (M1_final_x, M1_final_y, M1_final_z)

    def point_to_segment_distance(segment_start, segment_end, point):
        def distance_between_points(p1, p2):
            return math.sqrt((p2[0] - p1[0])**2 + (p2[1] - p1[1])**2 + (p2[2] - p1[2])**2)
        def dot_product(v1, v2):
            return v1[0] * v2[0] + v1[1] * v2[1] + v1[2] * v2[2]
        def vector_subtract(p1, p2):
            return (p1[0] - p2[0], p1[1] - p2[1], p1[2] - p2[2])

        AB = vector_subtract(segment_end, segment_start)
        AP = vector_subtract(point, segment_start)
        AB_squared = dot_product(AB, AB)
        if AB_squared == 0:
            return distance_between_points(point, segment_start)
        t = dot_product(AP, AB) / AB_squared
        if t < 0:
            return distance_between_points(point, segment_start)
        elif t > 1:
            return distance_between_points(point, segment_end)
        else:
            projection_point = (
                segment_start[0] + t * AB[0],
                segment_start[1] + t * AB[1],
                segment_start[2] + t * AB[2]
            )
            return distance_between_points(point, projection_point)

    real_point_1 = (7, 200, 0)
    real_point_2 = (-7, 200, 10)

    start_loop_time = smoke_center_start_time_total
    end_loop_time = smoke_center_start_time_total + smoke_valid_time + step

    t_valid_begin = 0
    t_valid_end = 0
    is_valid_started = False
    for t in np.arange(start_loop_time, end_loop_time, step):
        smoke_center_pos = calculate_smoke_center_position(t)
        M1_pos = calculate_M1_position(t)
        dist_1 = point_to_segment_distance(real_point_1, M1_pos, smoke_center_pos)
        dist_2 = point_to_segment_distance(real_point_2, M1_pos, smoke_center_pos)
        
        if dist_1 < 10 and dist_2 < 10 and not is_valid_started:
            t_valid_begin = t
            is_valid_started = True
        
        if (dist_1 >= 10 or dist_2 >= 10) and is_valid_started:
            t_valid_end = t
            break
    
    if is_valid_started and t_valid_end == 0:
        t_valid_end = end_loop_time

    return (t_valid_begin, t_valid_end), (new_x_drone, new_y_drone, new_z_drone), (smoke_bomb_final_x, smoke_bomb_final_y, smoke_bomb_final_z)

def set_all_seeds(seed):
    random.seed(seed)
    np.random.seed(seed)

def find_optimal_params_for_point(fyx_pos_init, t_bias=0, lb=None, ub=None, verbose=True):
    """
    针对给定的条件，使用遗传算法(GA)寻找最优的4个参数。

    参数:
    fyx_pos_init (tuple): 发烟信标的初始位置。
    t_bias (float): 时间偏差值。
    lb (list, optional): 4个参数的下界. 如果为None, 则使用默认值.
    ub (list, optional): 4个参数的上界. 如果为None, 则使用默认值.
    verbose (bool): 是否打印详细优化过程。

    返回:
    tuple: (best_params, max_duration)
    """
    if verbose:
        print("\n" + "="*50)
        print(f"开始为新条件优化: fyx_pos_init={fyx_pos_init}, t_bias={t_bias}")
        print("="*50)

    # --- 内部目标函数 (利用闭包) ---
    def schaffer(p):
        drone_speed, theta, t_drop, t_burst = p
        t_span, *_ = calculate_smoke_valid_duration(drone_speed, theta, t_drop, t_burst, 
                                                    fyx_pos_init=fyx_pos_init, 
                                                    t_bias=t_bias)
        duration = t_span[1] - t_span[0]
        if duration <= 0:
            return float('inf')
        return -duration

    # 如果用户没有传入lb, 则使用默认值
    if lb is None:
        lb = [70, 90, 0, 0]
    # 如果用户没有传入ub, 则使用默认值
    if ub is None:
        ub = [140, 180, 55, 55]
    
    if verbose:
        print(f"使用的搜索范围: lb={lb}, ub={ub}")

    # --- 运行遗传算法 ---
    if verbose:
        print("--- 开始运行遗传算法(GA) ---")
    
    # 将动态确定的lb和ub传入GA
    ga = GA(func=schaffer, n_dim=4, size_pop=100, max_iter=100, prob_mut=0.05, lb=lb, ub=ub, precision=1e-7)
    best_x, best_y = ga.run()
    
    # --- 处理并返回结果 ---
    final_best_params = best_x
    final_max_duration = -best_y[0] 
    
    if verbose:
        print("\n--- 优化完成 ---")
        print(f"找到的最佳参数: {np.round(final_best_params, 3)}")
        print(f"对应的最大有效时长: {final_max_duration:.4f} 秒")
        
    return final_best_params, final_max_duration

if __name__ == "__main__":
    
    def union_interval_length(intervals):
        """计算一系列时间区间的并集总长度"""
        valid_intervals = [iv for iv in intervals if iv[0] < iv[1]]
        if not valid_intervals:
            return 0
        
        sorted_intervals = sorted(valid_intervals, key=lambda x: x[0])
        
        merged = []
        for interval in sorted_intervals:
            if not merged or merged[-1][1] < interval[0]:
                merged.append(list(interval))
            else:
                merged[-1][1] = max(merged[-1][1], interval[1])
        
        total_length = sum(end - start for start, end in merged)
        return total_length

    SEED_VALUE = 666
    set_all_seeds(SEED_VALUE)

    # --- 第一次优化 ---
    fyx_pos_1 = (17800, 0, 1800)
    t_bias_1 = 0.0
    lb_1 = [70, 0, 0, 0]
    ub_1 = [140, 180, 5, 5]

    print(">>> 正在进行第 1 次优化...")
    best_params_1, max_duration_1 = find_optimal_params_for_point(
        fyx_pos_init=fyx_pos_1, t_bias=t_bias_1, lb=lb_1, ub=ub_1, verbose=False
    )
    
    print("\n" + "*"*20, "最终结果 1", "*"*20)
    print(f"对于 fyx_pos_init={fyx_pos_1}, t_bias={t_bias_1:.2f}:")
    print(f"  - 最佳参数: {np.round(best_params_1, 3)}")
    print(f"  - 最大时长: {max_duration_1:.4f} 秒")

    # --- 第二次优化 ---
    t_span_1_local, fyx_pos_2, burst_pos_1 = calculate_smoke_valid_duration(*best_params_1, fyx_pos_init=fyx_pos_1, t_bias=t_bias_1)
    t_bias_2 = best_params_1[2]  # 第二次事件的开始时间
    
    lb_2 = [70, 90, 1.5, 0]
    ub_2 = [140, 270, 5, 5]
    
    print("\n>>> 正在进行第 2 次优化...")
    best_params_2, max_duration_2 = find_optimal_params_for_point(
        fyx_pos_init=fyx_pos_2, t_bias=t_bias_2, lb=lb_2, ub=ub_2, verbose=False
    )

    print("\n" + "*"*20, "最终结果 2", "*"*20)
    print(f"对于 fyx_pos_init={fyx_pos_2}, t_bias={t_bias_2:.2f}:")
    print(f"  - 最佳参数: {np.round(best_params_2, 3)}")
    print(f"  - 最大时长: {max_duration_2:.4f} 秒")

    # --- 第三次优化 ---
    t_span_2_local, fyx_pos_3, burst_pos_2 = calculate_smoke_valid_duration(*best_params_2, fyx_pos_init=fyx_pos_2, t_bias=t_bias_2)
    t_bias_3 = t_bias_2 + best_params_2[2] # 第三次事件的开始时间
    
    lb_3 = [70, 90, 2.5, 0]
    ub_3 = [140, 270, 10, 15]

    print("\n>>> 正在进行第 3 次优化...")
    best_params_3, max_duration_3 = find_optimal_params_for_point(
        fyx_pos_init=fyx_pos_3, t_bias=t_bias_3, lb=lb_3, ub=ub_3, verbose=False
    )
    
    print("\n" + "*"*20, "最终结果 3", "*"*20)
    print(f"对于 fyx_pos_init={fyx_pos_3}, t_bias={t_bias_3:.2f}:")
    print(f"  - 最佳参数: {np.round(best_params_3, 3)}")
    print(f"  - 最大时长: {max_duration_3:.4f} 秒")

    # 在第三次优化后，计算它的本地有效时间区间
    t_span_3_local, fyx_pos_4, burst_pos_3 = calculate_smoke_valid_duration(*best_params_3, fyx_pos_init=fyx_pos_3, t_bias=t_bias_3)

    t_span_1_abs = (t_span_1_local[0] + t_bias_1, t_span_1_local[1] + t_bias_1)
    t_span_2_abs = (t_span_2_local[0] + t_bias_2, t_span_2_local[1] + t_bias_2)
    t_span_3_abs = (t_span_3_local[0] + t_bias_3, t_span_3_local[1] + t_bias_3)
    
    # --- 计算总有效时长 ---
    print("\n" + "="*20, "总有效时长计算", "="*20)
    
    # 收集所有的绝对时间区间
    all_absolute_spans = [t_span_1_abs, t_span_2_abs, t_span_3_abs]
    
    print("各次投放的有效时间区间 (绝对时间):")
    print(f"  第1次 (t_bias={t_bias_1:.2f}): 本地 {np.round(t_span_1_local, 2)} -> 绝对 {np.round(t_span_1_abs, 2)}")
    print(f"  第2次 (t_bias={t_bias_2:.2f}): 本地 {np.round(t_span_2_local, 2)} -> 绝对 {np.round(t_span_2_abs, 2)}")
    print(f"  第3次 (t_bias={t_bias_3:.2f}): 本地 {np.round(t_span_3_local, 2)} -> 绝对 {np.round(t_span_3_abs, 2)}")
    
    # 使用函数计算并集的总长度
    total_effective_duration = union_interval_length(all_absolute_spans)
    
    print("\n" + "-"*50)
    print(f"三次投放的总有效覆盖时长 (并集): {total_effective_duration:.4f} 秒")
    print("-"*50)

def calculate_smoke_valid_duration_2(drone_speed_1, theta_deg_1, t_drop_1, t_burst_1,
                                     drone_speed_2, theta_deg_2, t_drop_2, t_burst_2,
                                     drone_speed_3, theta_deg_3, t_drop_3, t_burst_3, fyx_pos_init_1=(17800,0,1800), step=0.01):
    t_duration_1, drop_pos_1, burst_pos_1 = calculate_smoke_valid_duration(drone_speed_1, theta_deg_1, t_drop_1, t_burst_1, fyx_pos_init_1, step=step)
    t_duration_2, drop_pos_2, burst_pos_2 = calculate_smoke_valid_duration(drone_speed_2, theta_deg_2, t_drop_2, t_burst_2, drop_pos_1, step=step, t_bias=t_drop_1)
    t_duration_3, drop_pos_3, burst_pos_3 = calculate_smoke_valid_duration(drone_speed_3, theta_deg_3, t_drop_3, t_burst_3, drop_pos_2, step=step, t_bias=t_drop_1 + t_drop_2)

    t_span_1 = [t_duration_1[0], t_duration_1[1]]
    t_span_2 = [t_duration_2[0] + t_drop_1, t_duration_2[1] + t_drop_1]
    t_span_3 = [t_duration_3[0] + t_drop_1 + t_drop_2, t_duration_3[1] + t_drop_1 + t_drop_2]
    def union_interval_length(intervals):
        # intervals: [(start1, end1), (start2, end2), ...]
        # 先按起点排序
        intervals = sorted(intervals, key=lambda x: x[0])
        merged = []
        for interval in intervals:
            if not merged or merged[-1][1] < interval[0]:
                merged.append(list(interval))
            else:
                merged[-1][1] = max(merged[-1][1], interval[1])
        # 计算总长度
        total = sum(end - start for start, end in merged)
        return total
    
    intervals = [t_span_1, t_span_2, t_span_3]
    total_duration = union_interval_length(intervals)
    # print(f"三者并集长度为: {total_duration:.2f}")

    t_1 = t_span_1[1] - t_span_1[0]
    t_2 = t_span_2[1] - t_span_2[0]
    t_3 = t_span_3[1] - t_span_3[0]
    t_spans = (t_1, t_2, t_3)
    smoke_inf_1 = (drop_pos_1, burst_pos_1, t_1)
    smoke_inf_2 = (drop_pos_2, burst_pos_2, t_2)
    smoke_inf_3 = (drop_pos_3, burst_pos_3, t_3)

    return  total_duration, smoke_inf_1, smoke_inf_2, smoke_inf_3, t_spans

def demo_func(params):
    total_duration, *_ = calculate_smoke_valid_duration_2(*params)
    return - total_duration

from sko.SA import SA
# --- 1. 定义边界约束 ---
gbest_x = np.concatenate((best_params_1, best_params_2, best_params_3))
lb = [70, 90, 0, 0, 70, 0, 1, 0, 70, 0, 1, 0]
ub = [140, 180, 5, 5, 140, 360, 5, 5, 140, 360, 5, 5]

# --- 2. 在初始化SA时传入lb和ub ---
sa = SA(func=demo_func, 
        x0=gbest_x, 
        T_max=10, 
        T_min=1e-9, 
        L=100, 
        max_stay_counter=150,
        lb=lb,
        ub=ub)

# --- 3. 运行算法并处理结果 (这部分代码保持不变) ---
best_x, best_y = sa.run()

# SA运行后
# 确保 best_x 是 numpy 数组或可迭代对象
formatted_sa_x = [round(float(x), 3) for x in best_x]
formatted_sa_y = round(float(-best_y), 2)  # 取反，保留两位小数
formatted_sa_y_hist = [-float(y) for y in sa.best_y_history]  # 全部取反

print('第三问：模拟退火算法最佳参数-值')
for i, x in enumerate(formatted_sa_x, 1):
    print(f'sa_best_x[{i}] = {x}', end=', ')
print('\nsa_best_y is', formatted_sa_y)


# t_1 = t_span_1_local[1] - t_span_1_local[0]
# t_2 = t_span_2_local[1] - t_span_2_local[0]
# t_3 = t_span_3_local[1] - t_span_3_local[0]
# smoke_inf_1 = (fyx_pos_2, burst_pos_1, t_1)
# smoke_inf_2 = (fyx_pos_3, burst_pos_2, t_2)
# smoke_inf_3 = (fyx_pos_4, burst_pos_3, t_3)
# t_spans = (t_1, t_2, t_3)
total_duration, smoke_inf_1, smoke_inf_2, smoke_inf_3, t_spans = calculate_smoke_valid_duration_2(*best_x)
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

excel_file_path = 'result1.xlsx'

def fmt(num, x):
    # 保留三位小数，如果不是数字则原样返回
    try:
        return round(float(num), x)
    except Exception:
        return num

data_to_fill = {
    '无人机运动方向': [fmt(best_params_1[1], 3), fmt(best_params_2[1], 3), fmt(best_params_3[1], 3)],
    '无人机运动速度 (m/s)': [fmt(best_params_1[0], 3), fmt(best_params_2[0], 3), fmt(best_params_3[0], 3)],
    '烟幕干扰弹投放点的x坐标 (m)': [fmt(smoke_inf_1[0][0], 3), fmt(smoke_inf_2[0][0], 3), fmt(smoke_inf_3[0][0], 3)],
    '烟幕干扰弹投放点的y坐标 (m)': [fmt(smoke_inf_1[0][1], 3), fmt(smoke_inf_2[0][1], 3), fmt(smoke_inf_3[0][1], 3)],
    '烟幕干扰弹投放点的z坐标 (m)': [fmt(smoke_inf_1[0][2], 3), fmt(smoke_inf_2[0][2], 3), fmt(smoke_inf_3[0][2], 3)],
    '烟幕干扰弹起爆点的x坐标 (m)': [fmt(smoke_inf_1[1][0], 3), fmt(smoke_inf_2[1][0], 3), fmt(smoke_inf_3[1][0], 3)],
    '烟幕干扰弹起爆点的y坐标 (m)': [fmt(smoke_inf_1[1][1], 3), fmt(smoke_inf_2[1][1], 3), fmt(smoke_inf_3[1][1], 3)],
    '烟幕干扰弹起爆点的z坐标 (m)': [fmt(smoke_inf_1[1][2], 3), fmt(smoke_inf_2[1][2], 3), fmt(smoke_inf_3[1][2], 3)],
    '有效干扰时长 (s)': [fmt(t_spans[0], 2), fmt(t_spans[1], 2), fmt(t_spans[2], 2)],
}

# 在保存Excel前加上用户确认
user_input = input("是否保存结果到Excel？输入Y或y保存，其他任意键不保存：")
if user_input.strip().lower() == 'y':
    workbook = load_workbook(excel_file_path)
    sheet = workbook.active

    # 获取列标题到列索引的映射
    column_map = {}
    for col_idx, cell in enumerate(sheet[1], 1):
        column_map[cell.value] = get_column_letter(col_idx)

    # 确定数据填充的起始行（第一行为标题行）
    start_row_index = 2 # Excel 行从 1 开始，标题占一行，数据从第 2 行开始

    # 准备要填充的数据，确保所有列表长度一致
    max_data_rows = 0
    for col_data in data_to_fill.values():
        max_data_rows = max(max_data_rows, len(col_data))

    for col_name in data_to_fill:
        while len(data_to_fill[col_name]) < max_data_rows:
            data_to_fill[col_name].append(None)

    # 填充数据
    for col_header, col_data in data_to_fill.items():
        if col_header in column_map:
            column_letter = column_map[col_header]
            for i, value in enumerate(col_data):
                target_row = start_row_index + i
                cell = sheet[f"{column_letter}{target_row}"]
                cell.value = value
        else:
            print(f"警告：列 '{col_header}' 在 Excel 文件中未找到，数据将不会填充。")
            
    # 保存修改后的 Excel 文件
    workbook.save(excel_file_path)
    print(f"成功使用新数据更新了 '{excel_file_path}'。")
else:
    print("未保存结果到Excel。")