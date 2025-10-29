import math
import numpy as np
import random
from sko.GA import GA
from sko.SA import SA

# ===================================================================
# 1. 核心物理模型与辅助函数 (此部分保持不变)
# ===================================================================

def set_all_seeds(seed):
    """设置全局随机种子以保证结果可复现。"""
    random.seed(seed)
    np.random.seed(seed)

def union_interval_length(intervals):
    """计算一系列时间区间的并集总长度。"""
    valid_intervals = [iv for iv in intervals if iv and iv[0] < iv[1]]
    if not valid_intervals:
        return 0
    
    sorted_intervals = sorted(valid_intervals, key=lambda x: x[0])
    
    merged = []
    for interval in sorted_intervals:
        if not merged or merged[-1][1] < interval[0]:
            merged.append(list(interval))
        else:
            merged[-1][1] = max(merged[-1][1], interval[1])
    
    total_length = sum(end - start for start, end in merged)
    return total_length

def calculate_smoke_valid_duration(drone_speed, theta_deg, t_drop, t_burst, 
                                   fyx_pos_init, M1_initial_point, 
                                   step=0.01, t_bias=0):
    """计算单次投放烟雾的有效遮蔽时长。"""
    # (物理模型代码与之前相同，为简洁省略)
    fake_point=(0, 0, 0); g=9.8; smoke_center_vz=-3; smoke_valid_time=20
    real_point_1 = (7, 200, 0); real_point_2 = (-7, 200, 10)
    vx = drone_speed * np.cos(np.radians(theta_deg)); vy = drone_speed * np.sin(np.radians(theta_deg))
    new_x_drone = fyx_pos_init[0] + vx * t_drop; new_y_drone = fyx_pos_init[1] + vy * t_drop; new_z_drone = fyx_pos_init[2]
    delta_z_smoke_bomb = 0.5 * (-g) * t_burst**2
    smoke_bomb_final_x = new_x_drone + vx * t_burst; smoke_bomb_final_y = new_y_drone + vy * t_burst; smoke_bomb_final_z = new_z_drone + delta_z_smoke_bomb
    smoke_center_start_time_local = t_drop + t_burst
    def get_smoke_pos(t_local):
        if t_local < smoke_center_start_time_local: return (smoke_bomb_final_x, smoke_bomb_final_y, smoke_bomb_final_z)
        dt = t_local - smoke_center_start_time_local; return (smoke_bomb_final_x, smoke_bomb_final_y, smoke_bomb_final_z + smoke_center_vz * dt)
    M1_direction_vector = np.array(fake_point) - np.array(M1_initial_point); M1_distance = np.linalg.norm(M1_direction_vector)
    M1_unit_vec = M1_direction_vector / M1_distance; M1_velocity_vec = 300 * M1_unit_vec
    def get_missile_pos(t_local):
        t_global = t_local + t_bias; return tuple(np.array(M1_initial_point) + M1_velocity_vec * t_global)
    def point_to_segment_distance(seg_start, seg_end, p):
        seg_start, seg_end, p = np.array(seg_start), np.array(seg_end), np.array(p)
        if np.array_equal(seg_start, seg_end): return np.linalg.norm(p - seg_start)
        line_vec, p_vec = seg_end - seg_start, p - seg_start
        t = np.dot(p_vec, line_vec) / np.dot(line_vec, line_vec)
        if t < 0.0: return np.linalg.norm(p - seg_start)
        if t > 1.0: return np.linalg.norm(p - seg_end)
        return np.linalg.norm(p - (seg_start + t * line_vec))
    t_valid_begin, t_valid_end = 0, 0; is_valid_started = False
    start_loop = smoke_center_start_time_local; end_loop = start_loop + smoke_valid_time
    for t in np.arange(start_loop, end_loop, step):
        smoke_pos = get_smoke_pos(t); missile_pos = get_missile_pos(t)
        dist_1 = point_to_segment_distance(real_point_1, missile_pos, smoke_pos); dist_2 = point_to_segment_distance(real_point_2, missile_pos, smoke_pos)
        is_obscured = dist_1 < 10 and dist_2 < 10
        if is_obscured and not is_valid_started: t_valid_begin = t; is_valid_started = True
        if not is_obscured and is_valid_started: t_valid_end = t; break
    if is_valid_started and t_valid_end == 0: t_valid_end = end_loop
    return (t_valid_begin, t_valid_end), (new_x_drone, new_y_drone, new_z_drone), (smoke_bomb_final_x, smoke_bomb_final_y, smoke_bomb_final_z)

# 使用ai进行函数封装，便于调用
# ===================================================================
# 2. 封装后的主优化函数 (已更新)
# ===================================================================
def optimize_smoke_strategy(initial_drone_pos, initial_missile_pos, 
                            custom_ranges=None, use_sa=True, verbose=True, seed=None):
    """
    为三次烟雾投放任务进行全流程优化，并允许自定义参数搜索范围。

    参数 (Args):
        initial_drone_pos (tuple): 无人机初始位置 (x, y, z)。
        initial_missile_pos (tuple): 导弹初始位置 (x, y, z)。
        custom_ranges (list, optional): 自定义参数搜索范围。
            应为一个包含12个子列表的列表, 每个子列表格式为 [min, max]。
            顺序为 [speed1, theta1, ..., t_burst3]。
            如果为 None, 则使用函数内置的默认范围。默认为 None。
        use_sa (bool): 是否执行模拟退火进行全局精调。默认为 True。
        verbose (bool): 是否打印详细的优化过程。默认为 True。
        seed (int, optional): 随机种子。默认为 None。

    返回 (Returns):
        dict: 包含最终优化结果的字典。
    """
    if seed is not None:
        set_all_seeds(seed)

    # --- 定义默认搜索范围 ---
    default_ga_ranges = {
        'drop1': {'lb': [70, 0, 0, 0], 'ub': [140, 360, 55, 55]},
        'drop2': {'lb': [70, 0, 1, 0], 'ub': [140, 360, 55, 55]},
        'drop3': {'lb': [70, 0, 1, 0], 'ub': [140, 360, 55, 55]}
    }
    default_sa_lb = [70, 0, 0, 0, 70, 0, 1, 0, 70, 0, 1, 0]
    default_sa_ub = [140, 360, 55, 55, 140, 360, 55, 55, 140, 360, 55, 55]

    # --- 根据输入选择使用默认范围还是自定义范围 ---
    if custom_ranges is None:
        if verbose: print("--- 使用默认参数搜索范围 ---")
        ga_ranges = default_ga_ranges
        sa_lb, sa_ub = default_sa_lb, default_sa_ub
    else:
        if verbose: print("--- 使用自定义参数搜索范围 ---")
        assert isinstance(custom_ranges, list) and len(custom_ranges) == 12, \
            "custom_ranges必须是一个包含12个[min, max]子列表的列表。"
        ga_ranges = {
            'drop1': {'lb': [r[0] for r in custom_ranges[0:4]], 'ub': [r[1] for r in custom_ranges[0:4]]},
            'drop2': {'lb': [r[0] for r in custom_ranges[4:8]], 'ub': [r[1] for r in custom_ranges[4:8]]},
            'drop3': {'lb': [r[0] for r in custom_ranges[8:12]], 'ub': [r[1] for r in custom_ranges[8:12]]},
        }
        sa_lb = [r[0] for r in custom_ranges]
        sa_ub = [r[1] for r in custom_ranges]

    # --- 内部辅助函数：单次投放的GA优化器 ---
    def find_optimal_params_for_drop(fyx_pos, t_bias, lb, ub, drop_number):
        if verbose: print(f"\n>>> [GA] 正在优化第 {drop_number} 次投放...")
        def objective_func(p):
            t_span, *_ = calculate_smoke_valid_duration(*p, fyx_pos_init=fyx_pos, M1_initial_point=initial_missile_pos, t_bias=t_bias)
            return -(t_span[1] - t_span[0])
        ga = GA(func=objective_func, n_dim=4, size_pop=50, max_iter=50, prob_mut=0.05, lb=lb, ub=ub, precision=1e-7)
        best_params, _ = ga.run()
        return best_params

    # --- 内部辅助函数：评估12维参数 ---
    def evaluate_12_params(p):
        p1, p2, p3 = p[0:4], p[4:8], p[8:12]
        span1, drop1, burst1 = calculate_smoke_valid_duration(*p1, initial_drone_pos, initial_missile_pos, t_bias=0)
        span2, drop2, burst2 = calculate_smoke_valid_duration(*p2, drop1, initial_missile_pos, t_bias=p1[2])
        span3, drop3, burst3 = calculate_smoke_valid_duration(*p3, drop2, initial_missile_pos, t_bias=p1[2] + p2[2])
        
        abs_span1 = (span1[0], span1[1]); abs_span2 = (span2[0] + p1[2], span2[1] + p1[2]); abs_span3 = (span3[0] + p1[2] + p2[2], span3[1] + p1[2] + p2[2])
        total_len = union_interval_length([abs_span1, abs_span2, abs_span3])
        infos = [
            {'drop_pos': drop1, 'burst_pos': burst1, 'duration': (span1[1] - span1[0]), 'spans': span1},
            {'drop_pos': drop2, 'burst_pos': burst2, 'duration': (span2[1] - span2[0]), 'spans': span2},
            {'drop_pos': drop3, 'burst_pos': burst3, 'duration': (span3[1] - span3[0]), 'spans': span3}
        ]
        return total_len, infos

    # --- 阶段一：序贯遗传算法（GA） ---
    if verbose: print("="*60 + "\n阶段一：开始使用序贯遗传算法 (GA) 进行初步优化...\n" + "="*60)
    fyx_pos_1 = initial_drone_pos
    params_1 = find_optimal_params_for_drop(fyx_pos_1, 0.0, **ga_ranges['drop1'], drop_number=1)
    _, fyx_pos_2, _ = calculate_smoke_valid_duration(*params_1, fyx_pos_1, initial_missile_pos)
    params_2 = find_optimal_params_for_drop(fyx_pos_2, params_1[2], **ga_ranges['drop2'], drop_number=2)
    _, fyx_pos_3, _ = calculate_smoke_valid_duration(*params_2, fyx_pos_2, initial_missile_pos, t_bias=params_1[2])
    params_3 = find_optimal_params_for_drop(fyx_pos_3, params_1[2] + params_2[2], **ga_ranges['drop3'], drop_number=3)
    
    ga_initial_solution = np.concatenate([params_1, params_2, params_3])
    ga_total_duration, ga_smoke_infos = evaluate_12_params(ga_initial_solution)
    if verbose: print(f"\n--- GA 初步优化完成 ---\n  - 总有效时长: {ga_total_duration:.4f} 秒")

    # --- 阶段二：模拟退火（SA） ---
    if use_sa:
        if verbose: print("\n" + "="*60 + "\n阶段二：开始使用模拟退火 (SA) 进行全局精调...\n" + "="*60)
        def sa_objective_func(p): return -evaluate_12_params(p)[0]
        sa = SA(func=sa_objective_func, x0=ga_initial_solution, T_max=10, T_min=1e-9, L=50, max_stay_counter=100, lb=sa_lb, ub=sa_ub)
        sa_best_x, _ = sa.run()
        final_total_duration, final_smoke_infos = evaluate_12_params(sa_best_x)
        final_best_params = np.array_split(sa_best_x, 3)
        method_used = "GA + SA"
        if verbose: print(f"\n--- SA 全局精调完成 ---\n  - 最终总有效时长: {final_total_duration:.4f} 秒")
    else:
        if verbose: print("\n" + "="*60 + "\n阶段二：跳过模拟退火 (SA) 精调。\n" + "="*60)
        final_best_params = np.array_split(ga_initial_solution, 3)
        final_total_duration = ga_total_duration
        final_smoke_infos = ga_smoke_infos
        method_used = "GA Only"

    return {'final_params': final_best_params, 'final_smoke_infos': final_smoke_infos, 'final_total_duration': final_total_duration, 'optimization_method': method_used}



import math
class Missile:
    """一个精简的导弹类，用于存储ID和轨迹的起止点。"""
    def __init__(self, missile_id: str, initial_position: tuple, target_position: tuple):
        self.id = missile_id
        self.initial_position = tuple(initial_position)
        self.target_position = tuple(target_position)
        self.t_spans = []

    def add_t_span(self, t_begin: float, t_end: float):
        self.t_spans.append(sorted([t_begin, t_end]))

    def get_total_t_span_length(self) -> float:
        if not self.t_spans: return 0
        intervals = sorted(self.t_spans, key=lambda x: x[0])
        merged = [list(intervals[0])]
        for start, end in intervals[1:]:
            if merged[-1][1] < start: merged.append([start, end])
            else: merged[-1][1] = max(merged[-1][1], end)
        return sum(end - start for start, end in merged)

    def __repr__(self):
        return f"Missile(id='{self.id}', pos={self.initial_position}, target={self.target_position})"


class Drone:
    """一个精简的无人机类，用于存储ID、位置，并接收分配的导弹ID。"""
    def __init__(self, drone_id: str, initial_position: tuple):
        self.id = drone_id
        self.initial_position = tuple(initial_position)
        self.associated_missile_id = None  # 等待被分配

    def __repr__(self):
        return f"Drone(id='{self.id}', pos={self.initial_position}, assigned_to='{self.associated_missile_id}')"

def point_to_segment_distance(segment_start: tuple, segment_end: tuple, point: tuple) -> float:
    """计算3D空间中一个点到一个线段的最短距离。"""
    line_vec = tuple(segment_end[i] - segment_start[i] for i in range(3))
    point_vec = tuple(point[i] - segment_start[i] for i in range(3))
    
    line_len_sq = sum(v**2 for v in line_vec)
    if line_len_sq == 0.0:
        return math.sqrt(sum(v**2 for v in point_vec))
        
    dot = sum(point_vec[i] * line_vec[i] for i in range(3))
    t = max(0, min(1, dot / line_len_sq))
    
    closest_point = tuple(segment_start[i] + t * line_vec[i] for i in range(3))
    dist_sq = sum((point[i] - closest_point[i])**2 for i in range(3))
    
    return math.sqrt(dist_sq)

def assign_drones_to_missiles(drones: dict, missiles: dict):
    """遍历所有无人机，根据到导弹轨迹的最近距离为其分配目标。"""
    print("\n--- 正在为无人机分配目标... ---")
    for drone_id, drone in drones.items():
        min_distance = float('inf')
        closest_missile_id = None
        
        for missile_id, missile in missiles.items():
            distance = point_to_segment_distance(
                missile.initial_position, 
                missile.target_position, 
                drone.initial_position
            )
            
            if distance < min_distance:
                min_distance = distance
                closest_missile_id = missile.id
        
        drone.associated_missile_id = closest_missile_id
        print(f"无人机 {drone.id} 分配给导弹 {closest_missile_id} (距离: {min_distance:.2f} 米)")
# ===================================================================
# 3. 示例用法 (已更新)
# ===================================================================

if __name__ == "__main__":
    # 导弹分配
    missile_target = (0, 200, 0)
    missile_positions = {
        'M1': (20000, 0, 2000), 
        'M2': (19000, 600, 2100), 
        'M3': (18000, -600, 1900)
    }
    drone_positions = {
        'FY1': (17800, 0, 1800), 
        'FY2': (12000, 1400, 1400), 
        'FY3': (6000, -3000, 700), 
        'FY4': (11000, 2000, 1800), 
        'FY5': (13000, -2000, 1300)
    }
    
    # --- 2. 实例化对象 ---
    missiles = {name: Missile(name, pos, missile_target) for name, pos in missile_positions.items()}
    drones = {name: Drone(name, pos) for name, pos in drone_positions.items()}

    # print("--- 对象创建完成 (初始状态) ---")
    # for d in drones.values():
    #     print(d)

    # --- 3. 执行核心的分配逻辑 ---
    assign_drones_to_missiles(drones, missiles)


    # --- 定义场景初始条件 ---
    RANDOM_SEED = 666

    # --- Helper function for printing results ---
    def print_summary(results):
        print("\n" + "="*28, f"最终结果 ({results['optimization_method']})", "="*28)
        param_labels = ['无人机速度', '飞行角度(θ)', '飞行时长(t_drop)', '下落时长(t_burst)']
        
        for i in range(3):
            params = results['final_params'][i]
            info = results['final_smoke_infos'][i]
            print(f"\n--- 第 {i+1} 次投放最终策略 ---")
            for label, val in zip(param_labels, params):
                print(f"  - {label:<16}: {val:.3f}")
            print(f"  - {'投放点位置':<16}: {np.round(info['drop_pos'], 2)}")
            print(f"  - {'爆炸点位置':<16}: {np.round(info['burst_pos'], 2)}")
            print(f"  - {'单次有效时长':<16}: {info['duration']:.4f} s")
            
        print("\n" + "*"*75)
        print(f"最终总有效时长: {results['final_total_duration']:.4f} 秒")
        print("*"*75)

    best_params = []
    smoke_inf = []
    associated_missiles = []

    for d in range(len(drones)):
        num = drones[f'FY{d+1}'].associated_missile_id
        associated_missiles.append(num)

    idx = 1
    custom_ranges_for_fy1 = [
        # 烟幕弹1: 限制速度为低速，角度为前半圈，投放时间短
        [70.0, 140.0],   # drone_speed_1
        [0.0, 180.0],    # theta_1
        [0.0, 5.0],      # t_drop_1
        [0.0, 5.0],      # t_burst_1
        # 烟幕弹2: 限制速度为高速，角度为后半圈
        [70.0, 140.0],  # drone_speed_2
        [160.0, 200.0],  # theta_2
        [1.5, 5.0],      # t_drop_2
        [0.0, 5.0],      # t_burst_2
        # 烟幕弹3: 范围与默认相同
        [100.0, 130.0],   # drone_speed_3
        [160.0, 200.0],   # theta_3
        [2.5, 5.0],      # t_drop_3
        [0.0, 7.0],      # t_burst_3
    ]
    ga_only_results = optimize_smoke_strategy(
        initial_drone_pos=drones[f'FY{idx}'].initial_position,
        initial_missile_pos=missiles[drones[f'FY{idx}'].associated_missile_id].initial_position,
        custom_ranges=custom_ranges_for_fy1, # 传递自定义范围
        use_sa=False, # 指定不执行SA
        verbose=True,
        seed=RANDOM_SEED
    )
    print_summary(ga_only_results)
    t_spans = (ga_only_results['final_smoke_infos'][0]['spans'], ga_only_results['final_smoke_infos'][1]['spans'], ga_only_results['final_smoke_infos'][2]['spans'])
    missiles[drones[f'FY{idx}'].associated_missile_id].add_t_span(t_spans[0][0], t_spans[0][1])
    missiles[drones[f'FY{idx}'].associated_missile_id].add_t_span(t_spans[1][0], t_spans[1][1])
    missiles[drones[f'FY{idx}'].associated_missile_id].add_t_span(t_spans[2][0], t_spans[2][1])
    best_params.append(ga_only_results['final_params'])
    smoke_inf.append([ga_only_results['final_smoke_infos'][0], ga_only_results['final_smoke_infos'][1], ga_only_results['final_smoke_infos'][2]])
    
    idx = 2
    custom_ranges_for_fy2 = [
        # 烟幕弹1: 限制速度为低速，角度为前半圈，投放时间短
        [70.0, 140.0],   # drone_speed_1
        [180.0, 360.0],    # theta_1
        [0.0, 5.0],      # t_drop_1
        [0.0, 5.0],      # t_burst_1
        # 烟幕弹2: 限制速度为高速，角度为后半圈
        [70.0, 140.0],  # drone_speed_2
        [90.0, 270.0],  # theta_2
        [1.0, 15.0],      # t_drop_2
        [0.0, 15.0],      # t_burst_2
        # 烟幕弹3: 范围与默认相同
        [70.0, 140.0],   # drone_speed_3
        [90.0, 270.0],   # theta_3
        [1.0, 15.0],      # t_drop_3
        [0.0, 15.0],      # t_burst_3
    ]
    ga_only_results = optimize_smoke_strategy(
        initial_drone_pos=drones[f'FY{idx}'].initial_position,
        initial_missile_pos=missiles[drones[f'FY{idx}'].associated_missile_id].initial_position,
        custom_ranges=custom_ranges_for_fy2, # 传递自定义范围
        use_sa=False, # 指定不执行SA
        verbose=True,
        seed=RANDOM_SEED
    )
    print_summary(ga_only_results)
    t_spans = (ga_only_results['final_smoke_infos'][0]['spans'], ga_only_results['final_smoke_infos'][1]['spans'], ga_only_results['final_smoke_infos'][2]['spans'])
    missiles[drones[f'FY{idx}'].associated_missile_id].add_t_span(t_spans[0][0], t_spans[0][1])
    missiles[drones[f'FY{idx}'].associated_missile_id].add_t_span(t_spans[1][0], t_spans[1][1])
    missiles[drones[f'FY{idx}'].associated_missile_id].add_t_span(t_spans[2][0], t_spans[2][1])
    best_params.append(ga_only_results['final_params'])
    smoke_inf.append([ga_only_results['final_smoke_infos'][0], ga_only_results['final_smoke_infos'][1], ga_only_results['final_smoke_infos'][2]])
 
    idx = 3
    custom_ranges_for_fy3 = [
        # 烟幕弹1: 限制速度为低速，角度为前半圈，投放时间短
        [90.0, 120.0],   # drone_speed_1
        [80.0, 100.0],    # theta_1
        [20.0, 30.0],      # t_drop_1
        [0.0, 5.0],      # t_burst_1
        # 烟幕弹2: 限制速度为高速，角度为后半圈
        [70.0, 140.0],  # drone_speed_2
        [0.0, 360.0],  # theta_2
        [1.0, 5.0],      # t_drop_2
        [0.0, 5.0],      # t_burst_2
        # 烟幕弹3: 范围与默认相同
        [70.0, 140.0],   # drone_speed_3
        [0.0, 360.0],   # theta_3
        [1.0, 5.0],      # t_drop_3
        [0.0, 5.0],      # t_burst_3
    ]
    ga_only_results = optimize_smoke_strategy(
        initial_drone_pos=drones[f'FY{idx}'].initial_position,
        initial_missile_pos=missiles[drones[f'FY{idx}'].associated_missile_id].initial_position,
        custom_ranges=custom_ranges_for_fy3, # 传递自定义范围
        use_sa=False,
        verbose=True,
        seed=RANDOM_SEED
    )
    print_summary(ga_only_results)
    t_spans = (ga_only_results['final_smoke_infos'][0]['spans'], ga_only_results['final_smoke_infos'][1]['spans'], ga_only_results['final_smoke_infos'][2]['spans'])
    missiles[drones[f'FY{idx}'].associated_missile_id].add_t_span(t_spans[0][0], t_spans[0][1])
    missiles[drones[f'FY{idx}'].associated_missile_id].add_t_span(t_spans[1][0], t_spans[1][1])
    missiles[drones[f'FY{idx}'].associated_missile_id].add_t_span(t_spans[2][0], t_spans[2][1])
    best_params.append(ga_only_results['final_params'])
    smoke_inf.append([ga_only_results['final_smoke_infos'][0], ga_only_results['final_smoke_infos'][1], ga_only_results['final_smoke_infos'][2]])
    
    idx = 4
    custom_ranges_for_fy4 = [
        # 烟幕弹1: 限制速度为低速，角度为前半圈，投放时间短
        [80.0, 110.0],   # drone_speed_1
        [200.0, 260.0],    # theta_1
        [5.0, 15.0],      # t_drop_1
        [5.0, 15.0],      # t_burst_1
        # 烟幕弹2: 限制速度为高速，角度为后半圈
        [70.0, 100.0],  # drone_speed_2
        [200.0, 230.0],  # theta_2
        [5.0, 15.0],      # t_drop_2
        [10.0, 15.0],      # t_burst_2
        # 烟幕弹3: 范围与默认相同
        [70.0, 140.0],   # drone_speed_3
        [90.0, 270.0],   # theta_3
        [1.0, 8.0],      # t_drop_3
        [1.0, 15.0],      # t_burst_3
    ]
    ga_only_results = optimize_smoke_strategy(
        initial_drone_pos=drones[f'FY{idx}'].initial_position,
        initial_missile_pos=missiles[drones[f'FY{idx}'].associated_missile_id].initial_position,
        custom_ranges=custom_ranges_for_fy4, # 传递自定义范围
        use_sa=False,
        verbose=True,
        seed=RANDOM_SEED
    )
    print_summary(ga_only_results)
    t_spans = (ga_only_results['final_smoke_infos'][0]['spans'], ga_only_results['final_smoke_infos'][1]['spans'], ga_only_results['final_smoke_infos'][2]['spans'])
    missiles[drones[f'FY{idx}'].associated_missile_id].add_t_span(t_spans[0][0], t_spans[0][1])
    missiles[drones[f'FY{idx}'].associated_missile_id].add_t_span(t_spans[1][0], t_spans[1][1])
    missiles[drones[f'FY{idx}'].associated_missile_id].add_t_span(t_spans[2][0], t_spans[2][1])
    best_params.append(ga_only_results['final_params'])
    smoke_inf.append([ga_only_results['final_smoke_infos'][0], ga_only_results['final_smoke_infos'][1], ga_only_results['final_smoke_infos'][2]])

    idx = 5
    custom_ranges_for_fy5 = [
        # 烟幕弹1: 限制速度为低速，角度为前半圈，投放时间短
        [90.0, 140.0],   # drone_speed_1
        [90.0, 180.0],    # theta_1
        [10.0, 20.0],      # t_drop_1
        [0.0, 5.0],      # t_burst_1
        # 烟幕弹2: 限制速度为高速，角度为后半圈
        [80.0, 120.0],  # drone_speed_2
        [80.0, 130.0],  # theta_2
        [1.0, 3.0],      # t_drop_2
        [0.0, 2.0],      # t_burst_2
        # 烟幕弹3: 范围与默认相同
        [70.0, 100.0],   # drone_speed_3
        [0.0, 90.0],   # theta_3
        [1.0, 4.0],      # t_drop_3
        [0.0, 1.0],      # t_burst_3
    ]
    ga_only_results = optimize_smoke_strategy(
        initial_drone_pos=drones[f'FY{idx}'].initial_position,
        initial_missile_pos=missiles[drones[f'FY{idx}'].associated_missile_id].initial_position,
        custom_ranges=custom_ranges_for_fy5, # 传递自定义范围
        use_sa=False,
        verbose=True,
        seed=RANDOM_SEED
    )
    print_summary(ga_only_results)
    t_spans = (ga_only_results['final_smoke_infos'][0]['spans'], ga_only_results['final_smoke_infos'][1]['spans'], ga_only_results['final_smoke_infos'][2]['spans'])
    missiles[drones[f'FY{idx}'].associated_missile_id].add_t_span(t_spans[0][0], t_spans[0][1])
    missiles[drones[f'FY{idx}'].associated_missile_id].add_t_span(t_spans[1][0], t_spans[1][1])
    missiles[drones[f'FY{idx}'].associated_missile_id].add_t_span(t_spans[2][0], t_spans[2][1])
    best_params.append(ga_only_results['final_params'])
    smoke_inf.append([ga_only_results['final_smoke_infos'][0], ga_only_results['final_smoke_infos'][1], ga_only_results['final_smoke_infos'][2]])
    
    t_M1 = missiles['M1'].get_total_t_span_length()
    t_M2 = missiles['M2'].get_total_t_span_length()
    t_M3 = missiles['M3'].get_total_t_span_length()
    print(f"各导弹遮挡总时长: {t_M1:.2f}, {t_M2:.2f}, {t_M3:.2f}")
    print(f"遮挡总时长: {t_M1 + t_M2 + t_M3:.2f}")
    # excel 输出
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    excel_file_path = 'result3.xlsx'

    def fmt(num, x):
        # 保留三位小数，如果不是数字则原样返回
        try:
            return round(float(num), x)
        except Exception:
            return num

    data_to_fill = {
        # '无人机运动方向': [fmt(best_params[0][1], 3), fmt(best_params[0][5], 3), fmt(best_params[0][9], 3), fmt(best_params[1][1], 3), fmt(best_params[1][5], 3), fmt(best_params[1][9], 3)],
        # '无人机运动速度 (m/s)': [fmt(best_params[0][0], 3), fmt(best_params[0][4], 3), fmt(best_params[0][8], 3), fmt(best_params[1][0], 3), fmt(best_params[1][4], 3), fmt(best_params[1][8], 3)],
        '无人机运动方向': [fmt(best_params[0][0][1], 3), fmt(best_params[0][1][1], 3), fmt(best_params[0][2][1], 3), 
                          fmt(best_params[1][0][1], 3), fmt(best_params[1][1][1], 3), fmt(best_params[1][2][1], 3),
                          fmt(best_params[2][0][1], 3), fmt(best_params[2][1][1], 3), fmt(best_params[2][2][1], 3),
                          fmt(best_params[3][0][1], 3), fmt(best_params[3][1][1], 3), fmt(best_params[3][2][1], 3),
                          fmt(best_params[4][0][1], 3), fmt(best_params[4][1][1], 3), fmt(best_params[4][2][1], 3),],
        '无人机运动速度 (m/s)': [fmt(best_params[0][0][0], 3), fmt(best_params[0][1][0], 3), fmt(best_params[0][2][0], 3),
                                fmt(best_params[1][0][0], 3), fmt(best_params[1][1][0], 3), fmt(best_params[1][2][0], 3),
                                fmt(best_params[2][0][0], 3), fmt(best_params[2][1][0], 3), fmt(best_params[2][2][0], 3),
                                fmt(best_params[3][0][0], 3), fmt(best_params[3][1][0], 3), fmt(best_params[3][2][0], 3),
                                fmt(best_params[4][0][0], 3), fmt(best_params[4][1][0], 3), fmt(best_params[4][2][0], 3)],
        '烟幕干扰弹投放点的x坐标 (m)': [fmt(smoke_inf[0][0]['drop_pos'][0], 3), fmt(smoke_inf[0][1]['drop_pos'][0], 3), fmt(smoke_inf[0][2]['drop_pos'][0], 3), 
                                       fmt(smoke_inf[1][0]['drop_pos'][0], 3), fmt(smoke_inf[1][1]['drop_pos'][0], 3), fmt(smoke_inf[1][2]['drop_pos'][0], 3),
                                       fmt(smoke_inf[2][0]['drop_pos'][0], 3), fmt(smoke_inf[2][1]['drop_pos'][0], 3), fmt(smoke_inf[2][2]['drop_pos'][0], 3),
                                       fmt(smoke_inf[3][0]['drop_pos'][0], 3), fmt(smoke_inf[3][1]['drop_pos'][0], 3), fmt(smoke_inf[3][2]['drop_pos'][0], 3),
                                       fmt(smoke_inf[4][0]['drop_pos'][0], 3), fmt(smoke_inf[4][1]['drop_pos'][0], 3), fmt(smoke_inf[4][2]['drop_pos'][0], 3)],
        '烟幕干扰弹投放点的y坐标 (m)': [fmt(smoke_inf[0][0]['drop_pos'][1], 3), fmt(smoke_inf[0][1]['drop_pos'][1], 3), fmt(smoke_inf[0][2]['drop_pos'][1], 3),
                                       fmt(smoke_inf[1][0]['drop_pos'][1], 3), fmt(smoke_inf[1][1]['drop_pos'][1], 3), fmt(smoke_inf[1][2]['drop_pos'][1], 3),
                                       fmt(smoke_inf[2][0]['drop_pos'][1], 3), fmt(smoke_inf[2][1]['drop_pos'][1], 3), fmt(smoke_inf[2][2]['drop_pos'][1], 3),
                                       fmt(smoke_inf[3][0]['drop_pos'][1], 3), fmt(smoke_inf[3][1]['drop_pos'][1], 3), fmt(smoke_inf[3][2]['drop_pos'][1], 3),
                                       fmt(smoke_inf[4][0]['drop_pos'][1], 3), fmt(smoke_inf[4][1]['drop_pos'][1], 3), fmt(smoke_inf[4][2]['drop_pos'][1], 3)],
        '烟幕干扰弹投放点的z坐标 (m)': [fmt(smoke_inf[0][0]['drop_pos'][2], 3), fmt(smoke_inf[0][1]['drop_pos'][2], 3), fmt(smoke_inf[0][2]['drop_pos'][2], 3),
                                       fmt(smoke_inf[1][0]['drop_pos'][2], 3), fmt(smoke_inf[1][1]['drop_pos'][2], 3), fmt(smoke_inf[1][2]['drop_pos'][2], 3),
                                       fmt(smoke_inf[2][0]['drop_pos'][2], 3), fmt(smoke_inf[2][1]['drop_pos'][2], 3), fmt(smoke_inf[2][2]['drop_pos'][2], 3),
                                       fmt(smoke_inf[3][0]['drop_pos'][2], 3), fmt(smoke_inf[3][1]['drop_pos'][2], 3), fmt(smoke_inf[3][2]['drop_pos'][2], 3),
                                       fmt(smoke_inf[4][0]['drop_pos'][2], 3), fmt(smoke_inf[4][1]['drop_pos'][2], 3), fmt(smoke_inf[4][2]['drop_pos'][2], 3)],
        '烟幕干扰弹起爆点的x坐标 (m)': [fmt(smoke_inf[0][0]['burst_pos'][0], 3), fmt(smoke_inf[0][1]['burst_pos'][0], 3), fmt(smoke_inf[0][2]['burst_pos'][0], 3),
                                      fmt(smoke_inf[1][0]['burst_pos'][0], 3), fmt(smoke_inf[1][1]['burst_pos'][0], 3), fmt(smoke_inf[1][2]['burst_pos'][0], 3),
                                      fmt(smoke_inf[2][0]['burst_pos'][0], 3), fmt(smoke_inf[2][1]['burst_pos'][0], 3), fmt(smoke_inf[2][2]['burst_pos'][0], 3),
                                      fmt(smoke_inf[3][0]['burst_pos'][0], 3), fmt(smoke_inf[3][1]['burst_pos'][0], 3), fmt(smoke_inf[3][2]['burst_pos'][0], 3),
                                      fmt(smoke_inf[4][0]['burst_pos'][0], 3), fmt(smoke_inf[4][1]['burst_pos'][0], 3), fmt(smoke_inf[4][2]['burst_pos'][0], 3)],
        '烟幕干扰弹起爆点的y坐标 (m)': [fmt(smoke_inf[0][0]['burst_pos'][1], 3), fmt(smoke_inf[0][1]['burst_pos'][1], 3), fmt(smoke_inf[0][2]['burst_pos'][1], 3),
                                      fmt(smoke_inf[1][0]['burst_pos'][1], 3), fmt(smoke_inf[1][1]['burst_pos'][1], 3), fmt(smoke_inf[1][2]['burst_pos'][1], 3),
                                      fmt(smoke_inf[2][0]['burst_pos'][1], 3), fmt(smoke_inf[2][1]['burst_pos'][1], 3), fmt(smoke_inf[2][2]['burst_pos'][1], 3),
                                      fmt(smoke_inf[3][0]['burst_pos'][1], 3), fmt(smoke_inf[3][1]['burst_pos'][1], 3), fmt(smoke_inf[3][2]['burst_pos'][1], 3),
                                      fmt(smoke_inf[4][0]['burst_pos'][1], 3), fmt(smoke_inf[4][1]['burst_pos'][1], 3), fmt(smoke_inf[4][2]['burst_pos'][1], 3)],
        '烟幕干扰弹起爆点的z坐标 (m)': [fmt(smoke_inf[0][0]['burst_pos'][2], 3), fmt(smoke_inf[0][1]['burst_pos'][2], 3), fmt(smoke_inf[0][2]['burst_pos'][2], 3),
                                      fmt(smoke_inf[1][0]['burst_pos'][2], 3), fmt(smoke_inf[1][1]['burst_pos'][2], 3), fmt(smoke_inf[1][2]['burst_pos'][2], 3),
                                      fmt(smoke_inf[2][0]['burst_pos'][2], 3), fmt(smoke_inf[2][1]['burst_pos'][2], 3), fmt(smoke_inf[2][2]['burst_pos'][2], 3),
                                      fmt(smoke_inf[3][0]['burst_pos'][2], 3), fmt(smoke_inf[3][1]['burst_pos'][2], 3), fmt(smoke_inf[3][2]['burst_pos'][2], 3),
                                      fmt(smoke_inf[4][0]['burst_pos'][2], 3), fmt(smoke_inf[4][1]['burst_pos'][2], 3), fmt(smoke_inf[4][2]['burst_pos'][2], 3)],
        '有效干扰时长 (s)': [fmt(smoke_inf[0][0]['duration'], 2), fmt(smoke_inf[0][1]['duration'], 2), fmt(smoke_inf[0][2]['duration'], 2),
                            fmt(smoke_inf[1][0]['duration'], 2), fmt(smoke_inf[1][1]['duration'], 2), fmt(smoke_inf[1][2]['duration'], 2),
                            fmt(smoke_inf[2][0]['duration'], 2), fmt(smoke_inf[2][1]['duration'], 2), fmt(smoke_inf[2][2]['duration'], 2),
                            fmt(smoke_inf[3][0]['duration'], 2), fmt(smoke_inf[3][1]['duration'], 2), fmt(smoke_inf[3][2]['duration'], 2),
                            fmt(smoke_inf[4][0]['duration'], 2), fmt(smoke_inf[4][1]['duration'], 2), fmt(smoke_inf[4][2]['duration'], 2)],
        '干扰的导弹编号' : [associated_missiles[0], associated_missiles[0], associated_missiles[0],
                         associated_missiles[1], associated_missiles[1], associated_missiles[1],
                          associated_missiles[2], associated_missiles[2], associated_missiles[2],
                          associated_missiles[3], associated_missiles[3], associated_missiles[3],
                          associated_missiles[4], associated_missiles[4], associated_missiles[4]]
    }

    # 在保存Excel前加上用户确认
    user_input = input("是否保存结果到Excel？输入Y或y保存，其他任意键不保存：")
    if user_input.strip().lower() == 'y':
        workbook = load_workbook(excel_file_path)
        sheet = workbook.active

        # 获取列标题到列索引的映射
        column_map = {}
        for col_idx, cell in enumerate(sheet[1], 1):
            column_map[cell.value] = get_column_letter(col_idx)

        # 确定数据填充的起始行（第一行为标题行）
        start_row_index = 2 # Excel 行从 1 开始，标题占一行，数据从第 2 行开始

        # 准备要填充的数据，确保所有列表长度一致
        max_data_rows = 0
        for col_data in data_to_fill.values():
            max_data_rows = max(max_data_rows, len(col_data))

        for col_name in data_to_fill:
            while len(data_to_fill[col_name]) < max_data_rows:
                data_to_fill[col_name].append(None)

        # 填充数据
        for col_header, col_data in data_to_fill.items():
            if col_header in column_map:
                column_letter = column_map[col_header]
                for i, value in enumerate(col_data):
                    target_row = start_row_index + i
                    cell = sheet[f"{column_letter}{target_row}"]
                    cell.value = value
            else:
                print(f"警告：列 '{col_header}' 在 Excel 文件中未找到，数据将不会填充。")
                
        # 保存修改后的 Excel 文件
        workbook.save(excel_file_path)
        print(f"成功使用新数据更新了 '{excel_file_path}'。")
    else:
        print("未保存结果到Excel。")
