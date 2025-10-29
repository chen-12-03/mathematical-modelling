import math
import numpy as np
import random

def calculate_smoke_valid_duration(drone_speed, theta_deg, t_drop, t_burst, fyx_pos_init, step=0.01, t_bias=0):
    # 常量定义
    fyx_point=fyx_pos_init
    fake_point=(0, 0, 0)
    g=9.8
    smoke_center_vz=-3
    smoke_valid_time=20

    # --- FY1飞行部分---
    vx = drone_speed * np.cos(np.radians(theta_deg))
    vy = drone_speed * np.sin(np.radians(theta_deg))
    vz = 0

    delta_x = vx * t_drop
    delta_y = vy * t_drop
    delta_z = vz * t_drop

    new_x_drone = fyx_point[0] + delta_x
    new_y_drone = fyx_point[1] + delta_y
    new_z_drone = fyx_point[2] + delta_z

    # --- 烟雾弹下落部分 ---
    smoke_bomb_initial_vx = vx
    smoke_bomb_initial_vy = vy
    smoke_bomb_initial_vz = vz

    delta_vz_gravity = -g * t_burst

    # smoke_bomb在z方向的最终速度
    smoke_bomb_vz_final = smoke_bomb_initial_vz + delta_vz_gravity

    # 计算smoke_bomb下落smoke_bomb_fall_time秒后的位置
    smoke_bomb_initial_x = new_x_drone
    smoke_bomb_initial_y = new_y_drone
    smoke_bomb_initial_z = new_z_drone

    delta_x_smoke_bomb = smoke_bomb_initial_vx * t_burst
    delta_y_smoke_bomb = smoke_bomb_initial_vy * t_burst
    delta_z_smoke_bomb = smoke_bomb_initial_vz * t_burst + 0.5 * (-g) * t_burst**2

    smoke_bomb_final_x = smoke_bomb_initial_x + delta_x_smoke_bomb
    smoke_bomb_final_y = smoke_bomb_initial_y + delta_y_smoke_bomb
    smoke_bomb_final_z = smoke_bomb_initial_z + delta_z_smoke_bomb

    # --- smoke_center运动部分 ---
    smoke_center_vx = 0
    smoke_center_vy = 0

    smoke_center_start_time_total = t_drop + t_burst

    def calculate_smoke_center_position(total_time_t):
        if total_time_t < smoke_center_start_time_total:
            return (smoke_bomb_final_x, smoke_bomb_final_y, smoke_bomb_final_z)
        else:
            smoke_center_motion_duration = total_time_t - smoke_center_start_time_total
            sc_x = smoke_bomb_final_x + smoke_center_vx * smoke_center_motion_duration
            sc_y = smoke_bomb_final_y + smoke_center_vy * smoke_center_motion_duration
            sc_z = smoke_bomb_final_z + smoke_center_vz * smoke_center_motion_duration
            return (sc_x, sc_y, sc_z)

    # --- M1运动部分 ---
    M1_initial_point = (20000, 0, 2000)
    M1_speed = 300

    M1_direction_vector = (
        fake_point[0] - M1_initial_point[0],
        fake_point[1] - M1_initial_point[1],
        fake_point[2] - M1_initial_point[2]
    )
    M1_distance = math.sqrt(
        M1_direction_vector[0]**2 + M1_direction_vector[1]**2 + M1_direction_vector[2]**2
    )
    M1_unit_direction_vector = (
        M1_direction_vector[0] / M1_distance,
        M1_direction_vector[1] / M1_distance,
        M1_direction_vector[2] / M1_distance
    )

    M1_vx = M1_speed * M1_unit_direction_vector[0]
    M1_vy = M1_speed * M1_unit_direction_vector[1]
    M1_vz = M1_speed * M1_unit_direction_vector[2]

    def calculate_M1_position(total_time_t):
        t_m1 = total_time_t + t_bias
        M1_delta_x = M1_vx * t_m1
        M1_delta_y = M1_vy * t_m1
        M1_delta_z = M1_vz * t_m1
        M1_final_x = M1_initial_point[0] + M1_delta_x
        M1_final_y = M1_initial_point[1] + M1_delta_y
        M1_final_z = M1_initial_point[2] + M1_delta_z
        return (M1_final_x, M1_final_y, M1_final_z)

    def point_to_segment_distance(segment_start, segment_end, point):
        def distance_between_points(p1, p2):
            return math.sqrt((p2[0] - p1[0])**2 + (p2[1] - p1[1])**2 + (p2[2] - p1[2])**2)
        def dot_product(v1, v2):
            return v1[0] * v2[0] + v1[1] * v2[1] + v1[2] * v2[2]
        def vector_subtract(p1, p2):
            return (p1[0] - p2[0], p1[1] - p2[1], p1[2] - p2[2])

        AB = vector_subtract(segment_end, segment_start)
        AP = vector_subtract(point, segment_start)
        AB_squared = dot_product(AB, AB)
        if AB_squared == 0:
            return distance_between_points(point, segment_start)
        t = dot_product(AP, AB) / AB_squared
        if t < 0:
            return distance_between_points(point, segment_start)
        elif t > 1:
            return distance_between_points(point, segment_end)
        else:
            projection_point = (
                segment_start[0] + t * AB[0],
                segment_start[1] + t * AB[1],
                segment_start[2] + t * AB[2]
            )
            return distance_between_points(point, projection_point)

    real_point_1 = (7, 200, 0)
    real_point_2 = (-7, 200, 10)

    start_loop_time = smoke_center_start_time_total
    end_loop_time = smoke_center_start_time_total + smoke_valid_time + step

    t_valid_begin = 0
    t_valid_end = 0
    for t in np.arange(start_loop_time, end_loop_time, step):
        smoke_center_pos = calculate_smoke_center_position(t)
        M1_pos = calculate_M1_position(t)
        dist_1 = point_to_segment_distance(real_point_1, M1_pos, smoke_center_pos)
        dist_2 = point_to_segment_distance(real_point_2, M1_pos, smoke_center_pos)
        if dist_1 < 10 and dist_2 < 10 and t_valid_begin == 0:
            t_valid_begin = t
        if dist_1 >=10 and dist_2 >= 10 and t_valid_begin != 0:
            t_valid_end = t
            break
    #      有效遮挡时间区间                 投放点坐标                              起爆点坐标
    return (t_valid_begin, t_valid_end), (new_x_drone, new_y_drone, new_z_drone), (smoke_bomb_final_x, smoke_bomb_final_y, smoke_bomb_final_z)

def calculate_smoke_valid_duration_2(drone_speed_1, theta_deg_1, t_drop_1, t_burst_1,
                                     drone_speed_2, theta_deg_2, t_drop_2, t_burst_2,
                                     drone_speed_3, theta_deg_3, t_drop_3, t_burst_3, fyx_pos_init_1=(17800,0,1800), step=0.01):
    t_duration_1, drop_pos_1, burst_pos_1 = calculate_smoke_valid_duration(drone_speed_1, theta_deg_1, t_drop_1, t_burst_1, fyx_pos_init_1, step=step)
    t_duration_2, drop_pos_2, burst_pos_2 = calculate_smoke_valid_duration(drone_speed_2, theta_deg_2, t_drop_2, t_burst_2, drop_pos_1, step=step, t_bias=t_drop_1)
    t_duration_3, drop_pos_3, burst_pos_3 = calculate_smoke_valid_duration(drone_speed_3, theta_deg_3, t_drop_3, t_burst_3, drop_pos_2, step=step, t_bias=t_drop_1 + t_drop_2)

    t_span_1 = [t_duration_1[0], t_duration_1[1]]
    t_span_2 = [t_duration_2[0] + t_drop_1, t_duration_2[1] + t_drop_1]
    t_span_3 = [t_duration_3[0] + t_drop_1 + t_drop_2, t_duration_3[1] + t_drop_1 + t_drop_2]
    def union_interval_length(intervals):
        # intervals: [(start1, end1), (start2, end2), ...]
        # 先按起点排序
        intervals = sorted(intervals, key=lambda x: x[0])
        merged = []
        for interval in intervals:
            if not merged or merged[-1][1] < interval[0]:
                merged.append(list(interval))
            else:
                merged[-1][1] = max(merged[-1][1], interval[1])
        # 计算总长度
        total = sum(end - start for start, end in merged)
        return total
    
    intervals = [t_span_1, t_span_2, t_span_3]
    total_duration = union_interval_length(intervals)
    # print(f"三者并集长度为: {total_duration:.2f}")

    t_1 = t_span_1[1] - t_span_1[0]
    t_2 = t_span_2[1] - t_span_2[0]
    t_3 = t_span_3[1] - t_span_3[0]
    t_spans = (t_1, t_2, t_3)
    smoke_inf_1 = (drop_pos_1, burst_pos_1, t_1)
    smoke_inf_2 = (drop_pos_2, burst_pos_2, t_2)
    smoke_inf_3 = (drop_pos_3, burst_pos_3, t_3)

    return  total_duration, smoke_inf_1, smoke_inf_2, smoke_inf_3, t_spans

def set_all_seeds(seed):
    random.seed(seed)
    np.random.seed(seed)

SEED_VALUE = 21
set_all_seeds(SEED_VALUE)

from sko.PSO import PSO
def demo_func(params):
    total_duration, *_ = calculate_smoke_valid_duration_2(*params)
    return - total_duration

pso = PSO(func=demo_func, dim=12, pop=50, max_iter=100, lb=[70, 90, 0, 0, 70, 0, 1, 0, 70, 0, 1, 0], ub=[140, 180, 5, 5, 140, 360, 10, 10, 140, 360, 10, 10], w=0.8, c1=0.5, c2=0.5)
pso.run()

formatted_x = [round(x, 3) for x in pso.gbest_x]
formatted_y = round(-pso.gbest_y, 2)
formatted_y_hist = [-y for y in pso.gbest_y_hist]

print('第三问：粒子群算法最佳参数-值')
for i, x in enumerate(formatted_x, 1):
    print(f'best_x[{i}] = {x}',end=', ')
print('best_y is', formatted_y)

from sko.SA import SA
# --- 1. 定义你的边界约束 ---
# 这是你提供的边界
lb = [70, 90, 0, 0, 70, 0, 1, 0, 70, 0, 1, 0]
ub = [140, 180, 5, 5, 140, 360, 5, 5, 140, 360, 5, 5]

# --- 2. 在初始化SA时传入lb和ub ---
# 假设 pso.gbest_x 和 demo_func 已经定义好了
# 我们在原来的参数列表后面加上 lb=lb, ub=ub
sa = SA(func=demo_func, 
        x0=pso.gbest_x, 
        T_max=10, 
        T_min=1e-9, 
        L=100, 
        max_stay_counter=150,
        lb=lb,  # <--- 添加下界约束
        ub=ub)  # <--- 添加上界约束

# --- 3. 运行算法并处理结果 (这部分代码保持不变) ---
best_x, best_y = sa.run()

# SA运行后
# 确保 best_x 是 numpy 数组或可迭代对象
formatted_sa_x = [round(float(x), 3) for x in best_x]
formatted_sa_y = round(float(-best_y), 2)  # 取反，保留两位小数
formatted_sa_y_hist = [-float(y) for y in sa.best_y_history]  # 全部取反

print('第三问：模拟退火算法最佳参数-值')
for i, x in enumerate(formatted_sa_x, 1):
    print(f'sa_best_x[{i}] = {x}', end=', ')
print('\nsa_best_y is', formatted_sa_y)

import matplotlib.pyplot as plt

fig, axs = plt.subplots(1, 2, figsize=(14, 5))  # 1行2列

axs[0].plot(formatted_y_hist)
axs[0].set_xlabel('Iteration')
axs[0].set_ylabel('Best Y')
axs[0].set_title('q3:PSO Optimization History')
axs[0].grid(True)

axs[1].plot(formatted_sa_y_hist)
axs[1].set_xlabel('Iteration')
axs[1].set_ylabel('Best Y')
axs[1].set_title('q3:SA Optimization History')
axs[1].grid(True)

plt.tight_layout()
plt.show()

total_duration, smoke_inf_1, smoke_inf_2, smoke_inf_3, t_spans = calculate_smoke_valid_duration_2(*best_x)

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

excel_file_path = 'result1.xlsx'

def fmt(num, x):
    # 保留三位小数，如果不是数字则原样返回
    try:
        return round(float(num), x)
    except Exception:
        return num

data_to_fill = {
    '无人机运动方向': [fmt(pso.gbest_x[1], 3), fmt(pso.gbest_x[5], 3), fmt(pso.gbest_x[9], 3)],
    '无人机运动速度 (m/s)': [fmt(pso.gbest_x[0], 3), fmt(pso.gbest_x[4], 3), fmt(pso.gbest_x[8], 3)],
    '烟幕干扰弹投放点的x坐标 (m)': [fmt(smoke_inf_1[0][0], 3), fmt(smoke_inf_2[0][0], 3), fmt(smoke_inf_3[0][0], 3)],
    '烟幕干扰弹投放点的y坐标 (m)': [fmt(smoke_inf_1[0][1], 3), fmt(smoke_inf_2[0][1], 3), fmt(smoke_inf_3[0][1], 3)],
    '烟幕干扰弹投放点的z坐标 (m)': [fmt(smoke_inf_1[0][2], 3), fmt(smoke_inf_2[0][2], 3), fmt(smoke_inf_3[0][2], 3)],
    '烟幕干扰弹起爆点的x坐标 (m)': [fmt(smoke_inf_1[1][0], 3), fmt(smoke_inf_2[1][0], 3), fmt(smoke_inf_3[1][0], 3)],
    '烟幕干扰弹起爆点的y坐标 (m)': [fmt(smoke_inf_1[1][1], 3), fmt(smoke_inf_2[1][1], 3), fmt(smoke_inf_3[1][1], 3)],
    '烟幕干扰弹起爆点的z坐标 (m)': [fmt(smoke_inf_1[1][2], 3), fmt(smoke_inf_2[1][2], 3), fmt(smoke_inf_3[1][2], 3)],
    '有效干扰时长 (s)': [fmt(t_spans[0], 2), fmt(t_spans[1], 2), fmt(t_spans[2], 2)],
}

# 在保存Excel前加上用户确认
user_input = input("是否保存结果到Excel？输入Y或y保存，其他任意键不保存：")
if user_input.strip().lower() == 'y':
    workbook = load_workbook(excel_file_path)
    sheet = workbook.active

    # 获取列标题到列索引的映射
    column_map = {}
    for col_idx, cell in enumerate(sheet[1], 1):
        column_map[cell.value] = get_column_letter(col_idx)

    # 确定数据填充的起始行（第一行为标题行）
    start_row_index = 2 # Excel 行从 1 开始，标题占一行，数据从第 2 行开始

    # 准备要填充的数据，确保所有列表长度一致
    max_data_rows = 0
    for col_data in data_to_fill.values():
        max_data_rows = max(max_data_rows, len(col_data))

    for col_name in data_to_fill:
        while len(data_to_fill[col_name]) < max_data_rows:
            data_to_fill[col_name].append(None)

    # 填充数据
    for col_header, col_data in data_to_fill.items():
        if col_header in column_map:
            column_letter = column_map[col_header]
            for i, value in enumerate(col_data):
                target_row = start_row_index + i
                cell = sheet[f"{column_letter}{target_row}"]
                cell.value = value
        else:
            print(f"警告：列 '{col_header}' 在 Excel 文件中未找到，数据将不会填充。")
            
    # 保存修改后的 Excel 文件
    workbook.save(excel_file_path)
    print(f"成功使用新数据更新了 '{excel_file_path}'。")
else:
    print("未保存结果到Excel。")