import math
import numpy as np
import random
import itertools

def calculate_smoke_valid_duration(drone_speed, theta_deg, t_drop, t_burst, fyx_pos_init, step=0.01, t_bias=0):
    # 常量定义
    fyx_point=fyx_pos_init
    fake_point=(0, 0, 0)
    g=9.8
    smoke_center_vz=-3
    smoke_valid_time=20

    # --- FY1飞行部分---
    vx = drone_speed * np.cos(np.radians(theta_deg))
    vy = drone_speed * np.sin(np.radians(theta_deg))
    vz = 0

    delta_x = vx * t_drop
    delta_y = vy * t_drop
    delta_z = vz * t_drop

    new_x_drone = fyx_point[0] + delta_x
    new_y_drone = fyx_point[1] + delta_y
    new_z_drone = fyx_point[2] + delta_z

    # --- 烟雾弹下落部分 ---
    smoke_bomb_initial_vx = vx
    smoke_bomb_initial_vy = vy
    smoke_bomb_initial_vz = vz

    delta_vz_gravity = -g * t_burst

    # smoke_bomb在z方向的最终速度
    smoke_bomb_vz_final = smoke_bomb_initial_vz + delta_vz_gravity

    # 计算smoke_bomb下落smoke_bomb_fall_time秒后的位置
    smoke_bomb_initial_x = new_x_drone
    smoke_bomb_initial_y = new_y_drone
    smoke_bomb_initial_z = new_z_drone

    delta_x_smoke_bomb = smoke_bomb_initial_vx * t_burst
    delta_y_smoke_bomb = smoke_bomb_initial_vy * t_burst
    delta_z_smoke_bomb = smoke_bomb_initial_vz * t_burst + 0.5 * (-g) * t_burst**2

    smoke_bomb_final_x = smoke_bomb_initial_x + delta_x_smoke_bomb
    smoke_bomb_final_y = smoke_bomb_initial_y + delta_y_smoke_bomb
    smoke_bomb_final_z = smoke_bomb_initial_z + delta_z_smoke_bomb

    # --- smoke_center运动部分 ---
    smoke_center_vx = 0
    smoke_center_vy = 0

    smoke_center_start_time_total = t_drop + t_burst

    def calculate_smoke_center_position(total_time_t):
        if total_time_t < smoke_center_start_time_total:
            return (smoke_bomb_final_x, smoke_bomb_final_y, smoke_bomb_final_z)
        else:
            smoke_center_motion_duration = total_time_t - smoke_center_start_time_total
            sc_x = smoke_bomb_final_x + smoke_center_vx * smoke_center_motion_duration
            sc_y = smoke_bomb_final_y + smoke_center_vy * smoke_center_motion_duration
            sc_z = smoke_bomb_final_z + smoke_center_vz * smoke_center_motion_duration
            return (sc_x, sc_y, sc_z)

    # --- M1运动部分 ---
    M1_initial_point = (20000, 0, 2000)
    M1_speed = 300

    M1_direction_vector = (
        fake_point[0] - M1_initial_point[0],
        fake_point[1] - M1_initial_point[1],
        fake_point[2] - M1_initial_point[2]
    )
    M1_distance = math.sqrt(
        M1_direction_vector[0]**2 + M1_direction_vector[1]**2 + M1_direction_vector[2]**2
    )
    M1_unit_direction_vector = (
        M1_direction_vector[0] / M1_distance,
        M1_direction_vector[1] / M1_distance,
        M1_direction_vector[2] / M1_distance
    )

    M1_vx = M1_speed * M1_unit_direction_vector[0]
    M1_vy = M1_speed * M1_unit_direction_vector[1]
    M1_vz = M1_speed * M1_unit_direction_vector[2]

    def calculate_M1_position(total_time_t):
        t_m1 = total_time_t + t_bias
        M1_delta_x = M1_vx * t_m1
        M1_delta_y = M1_vy * t_m1
        M1_delta_z = M1_vz * t_m1
        M1_final_x = M1_initial_point[0] + M1_delta_x
        M1_final_y = M1_initial_point[1] + M1_delta_y
        M1_final_z = M1_initial_point[2] + M1_delta_z
        return (M1_final_x, M1_final_y, M1_final_z)

    def point_to_segment_distance(segment_start, segment_end, point):
        def distance_between_points(p1, p2):
            return math.sqrt((p2[0] - p1[0])**2 + (p2[1] - p1[1])**2 + (p2[2] - p1[2])**2)
        def dot_product(v1, v2):
            return v1[0] * v2[0] + v1[1] * v2[1] + v1[2] * v2[2]
        def vector_subtract(p1, p2):
            return (p1[0] - p2[0], p1[1] - p2[1], p1[2] - p2[2])

        AB = vector_subtract(segment_end, segment_start)
        AP = vector_subtract(point, segment_start)
        AB_squared = dot_product(AB, AB)
        if AB_squared == 0:
            return distance_between_points(point, segment_start)
        t = dot_product(AP, AB) / AB_squared
        if t < 0:
            return distance_between_points(point, segment_start)
        elif t > 1:
            return distance_between_points(point, segment_end)
        else:
            projection_point = (
                segment_start[0] + t * AB[0],
                segment_start[1] + t * AB[1],
                segment_start[2] + t * AB[2]
            )
            return distance_between_points(point, projection_point)

    real_point_1 = (7, 200, 0)
    real_point_2 = (-7, 200, 10)

    start_loop_time = smoke_center_start_time_total
    end_loop_time = smoke_center_start_time_total + smoke_valid_time + step

    t_valid_begin = 0
    t_valid_end = 0
    for t in np.arange(start_loop_time, end_loop_time, step):
        smoke_center_pos = calculate_smoke_center_position(t)
        M1_pos = calculate_M1_position(t)
        dist_1 = point_to_segment_distance(real_point_1, M1_pos, smoke_center_pos)
        dist_2 = point_to_segment_distance(real_point_2, M1_pos, smoke_center_pos)
        if dist_1 < 10 and dist_2 < 10 and t_valid_begin == 0:
            t_valid_begin = t
        if dist_1 >=10 and dist_2 >= 10 and t_valid_begin != 0:
            t_valid_end = t
            break
    #      有效遮挡时间区间                 投放点坐标                              起爆点坐标
    return (t_valid_begin, t_valid_end), (new_x_drone, new_y_drone, new_z_drone), (smoke_bomb_final_x, smoke_bomb_final_y, smoke_bomb_final_z)

def calculate_smoke_valid_duration_3(drone_speed_1, theta_deg_1, t_drop_1, t_burst_1,
                                     drone_speed_2, theta_deg_2, t_drop_2, t_burst_2,
                                     drone_speed_3, theta_deg_3, t_drop_3, t_burst_3, 
                                     fyx_pos_init_1, fyx_pos_init_2, fyx_pos_init_3, step=0.01):
    t_duration_1, drop_pos_1, burst_pos_1 = calculate_smoke_valid_duration(drone_speed_1, theta_deg_1, t_drop_1, t_burst_1, fyx_pos_init_1, step=step)
    t_duration_2, drop_pos_2, burst_pos_2 = calculate_smoke_valid_duration(drone_speed_2, theta_deg_2, t_drop_2, t_burst_2, fyx_pos_init_2, step=step)
    t_duration_3, drop_pos_3, burst_pos_3 = calculate_smoke_valid_duration(drone_speed_3, theta_deg_3, t_drop_3, t_burst_3, fyx_pos_init_3, step=step)

    t_span_1 = [t_duration_1[0], t_duration_1[1]]
    t_span_2 = [t_duration_2[0], t_duration_2[1]]
    t_span_3 = [t_duration_3[0], t_duration_3[1]]
    def union_interval_length(intervals):
        # intervals: [(start1, end1), (start2, end2), ...]
        # 先按起点排序
        intervals = sorted(intervals, key=lambda x: x[0])
        merged = []
        for interval in intervals:
            if not merged or merged[-1][1] < interval[0]:
                merged.append(list(interval))
            else:
                merged[-1][1] = max(merged[-1][1], interval[1])
        # 计算总长度
        total = sum(end - start for start, end in merged)
        return total
    
    intervals = [t_span_1, t_span_2, t_span_3]
    total_duration = union_interval_length(intervals)

    t_1 = t_span_1[1] - t_span_1[0]
    t_2 = t_span_2[1] - t_span_2[0]
    t_3 = t_span_3[1] - t_span_3[0]
    t_spans = (t_1, t_2, t_3)
    smoke_inf_1 = (drop_pos_1, burst_pos_1, t_1)
    smoke_inf_2 = (drop_pos_2, burst_pos_2, t_2)
    smoke_inf_3 = (drop_pos_3, burst_pos_3, t_3)

    return  total_duration, smoke_inf_1, smoke_inf_2, smoke_inf_3, t_spans

# 遗传算法实现
class Population:
    """
    遗传算法种群类。
    针对12个优化变量 (3组drone_speed, theta, t_drop, t_burst) 进行优化。
    """
    def __init__(self, size, chrom_size, cp, mp, gen_max, fyx_point_1, fyx_point_2, fyx_point_3):
        # 种群信息
        self.individuals = []          # 个体集合, 每个个体有12个染色体
        self.fitness = []              # 个体适应度集
        self.selector_probability = [] # 个体选择概率集合
        self.new_individuals = []      # 新一代个体集合

        # 精英个体信息
        self.elitist = {'chromosome': [0]*12, 'fitness': 0.0, 'age': 0}

        self.size = size
        self.chromosome_size = chrom_size
        self.crossover_probability = cp
        self.mutation_probability = mp
        self.generation_max = gen_max
        self.age = 0

        self.fyx_point_1 = fyx_point_1
        self.fyx_point_2 = fyx_point_2
        self.fyx_point_3 = fyx_point_3

        # 定义每个优化变量的取值范围
        # 3组，每组4个参数
        self.variable_ranges = [
            [70.0, 140.0],   # drone_speed_1
            [0.0, 360.0],    # theta_1
            [0.0, 5.0],     # t_drop_1
            [0.0, 5.0],     # t_burst_1
            [70.0, 140.0],   # drone_speed_2
            [0.0, 360.0],    # theta_2
            [10.0, 50.0],     # t_drop_2
            [0.0, 15.0],     # t_burst_2
            [70.0, 140.0],   # drone_speed_3
            [0.0, 180.0],    # theta_3
            [30.0, 90.0],     # t_drop_3
            [0.0, 10.0],     # t_burst_3
        ]

        max_chrom_value = 2 ** self.chromosome_size - 1
        for _ in range(self.size):
            self.individuals.append([random.randint(0, max_chrom_value) for _ in range(12)])
            self.new_individuals.append([0]*12)
            self.fitness.append(0)
            self.selector_probability.append(0)

    def decode(self, interval, chromosome):
        d = interval[1] - interval[0]
        max_chrom_value = float(2 ** self.chromosome_size - 1)
        return interval[0] + chromosome * d / max_chrom_value

    def fitness_func(self, individual_chromosomes):
        # 解码12个参数
        params = [self.decode(self.variable_ranges[i], individual_chromosomes[i]) for i in range(12)]
        # 适应度为三段烟雾并集时长
        total_duration, *_ = calculate_smoke_valid_duration_3(
            params[0], params[1], params[2], params[3],
            params[4], params[5], params[6], params[7],
            params[8], params[9], params[10], params[11],
            self.fyx_point_1, self.fyx_point_2, self.fyx_point_3  # 传递初始点
        )
        return total_duration

    def evaluate(self):
        for i in range(self.size):
            self.fitness[i] = self.fitness_func(self.individuals[i])
        ft_sum = sum(self.fitness)
        if ft_sum == 0:
            for i in range(self.size):
                self.selector_probability[i] = 1.0 / self.size
        else:
            for i in range(self.size):
                self.selector_probability[i] = self.fitness[i] / ft_sum
        for i in range(1, self.size):
            self.selector_probability[i] += self.selector_probability[i-1]

    def select(self):
        rand_prob = random.random()
        for i, p in enumerate(self.selector_probability):
            if p > rand_prob:
                return i
        return self.size - 1

    def cross(self, chrom1, chrom2):
        if chrom1 != chrom2 and random.random() < self.crossover_probability:
            cross_point = random.randint(1, self.chromosome_size - 1)
            mask = (2 ** self.chromosome_size - 1) << cross_point
            part1_c1 = chrom1 & mask
            part2_c1 = chrom1 & ~mask
            part1_c2 = chrom2 & mask
            part2_c2 = chrom2 & ~mask
            new_chrom1 = part1_c1 | part2_c2
            new_chrom2 = part1_c2 | part2_c1
            return new_chrom1, new_chrom2
        return chrom1, chrom2

    def mutate(self, chrom):
        if random.random() < self.mutation_probability:
            mutation_point = random.randint(1, self.chromosome_size)
            mask = 1 << (mutation_point - 1)
            chrom ^= mask
        return chrom

    def reproduct_elitist(self):
        current_best_idx = self.fitness.index(max(self.fitness))
        if self.elitist['fitness'] < self.fitness[current_best_idx]:
            self.elitist['fitness'] = self.fitness[current_best_idx]
            self.elitist['chromosome'] = self.individuals[current_best_idx][:]
            self.elitist['age'] = self.age

    def evolve(self):
        self.evaluate()
        self.reproduct_elitist()
        i = 0
        while i < self.size:
            parent1_idx = self.select()
            parent2_idx = self.select()
            parent1 = self.individuals[parent1_idx]
            parent2 = self.individuals[parent2_idx]
            child1 = [0] * 12
            child2 = [0] * 12
            for j in range(12):
                (c1, c2) = self.cross(parent1[j], parent2[j])
                child1[j] = self.mutate(c1)
                child2[j] = self.mutate(c2)
            self.new_individuals[i] = child1
            if i + 1 < self.size:
                self.new_individuals[i+1] = child2
            i += 2
        self.individuals = self.new_individuals[:]
        new_fitness = [self.fitness_func(ind) for ind in self.individuals]
        worst_idx = new_fitness.index(min(new_fitness))
        self.individuals[worst_idx] = self.elitist['chromosome'][:]

    def run(self):
        for i in range(self.generation_max):
            self.age = i
            self.evolve()
            print(f"世代 {i:3d}/{self.generation_max}: 最佳时长 = {max(self.fitness):.2f}, "
                  f"平均时长 = {sum(self.fitness)/self.size:.2f}, "
                  f"历史最佳 = {self.elitist['fitness']:.2f}")
        print("                 优化结束                 ")
        best_chromosomes = self.elitist['chromosome']
        best_params = [self.decode(self.variable_ranges[j], best_chromosomes[j]) for j in range(12)]
        print("最优参数：")
        for k in range(3):
            print(f"  drone{k+1}_speed: {best_params[4*k]:.4f} m/s")
            print(f"  theta{k+1}:      {best_params[4*k+1]:.4f} degrees")
            print(f"  t_drop{k+1}:     {best_params[4*k+2]:.4f} s")
            print(f"  t_burst{k+1}:    {best_params[4*k+3]:.4f} s")
        print(f"\n最佳遮蔽时长: {self.elitist['fitness']:.2f}")

        # 返回最优参数
        return best_params

# --- 遗传算法参数配置 ---
POPULATION_SIZE = 200       # 种群大小
CHROMOSOME_BITS = 20       # 单个染色体的位数（精度）
CROSSOVER_PROB = 0.8       # 交叉概率
MUTATION_PROB = 0.10       # 变异概率
MAX_GENERATIONS = 75       # 最大进化代数

# 无人机初始位置定义
fy1_point = (17800, 0, 1800)
fy2_point = (12000, 1400, 1400)
fy3_point = (6000, -3000, 700)
fy4_point = (11000, 2000, 1800)
fy5_point = (13000, -2000, 1300)
fyx_points = [fy1_point, fy2_point, fy3_point, fy4_point, fy5_point]

print("开始遗传优化算法...")
pop = Population(
    size=POPULATION_SIZE, 
    chrom_size=CHROMOSOME_BITS, 
    cp=CROSSOVER_PROB, 
    mp=MUTATION_PROB, 
    gen_max=MAX_GENERATIONS,
    fyx_point_1=fyx_points[0],
    fyx_point_2=fyx_points[1],
    fyx_point_3=fyx_points[2]
)

# 调用最优参数，获取关键数据
best_params = pop.run()
total_duration, smoke_inf_1, smoke_inf_2, smoke_inf_3, t_spans = calculate_smoke_valid_duration_3(*best_params, fyx_points[0], fyx_points[1], fyx_points[2])

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

excel_file_path = 'result2.xlsx'

def fmt(num, x):
    # 保留三位小数，如果不是数字则原样返回
    try:
        return round(float(num), x)
    except Exception:
        return num

data_to_fill = {
    '无人机运动方向': [fmt(best_params[1], 3), fmt(best_params[5], 3), fmt(best_params[9], 3)],
    '无人机运动速度 (m/s)': [fmt(best_params[0], 3), fmt(best_params[4], 3), fmt(best_params[8], 3)],
    '烟幕干扰弹投放点的x坐标 (m)': [fmt(smoke_inf_1[0][0], 3), fmt(smoke_inf_2[0][0], 3), fmt(smoke_inf_3[0][0], 3)],
    '烟幕干扰弹投放点的y坐标 (m)': [fmt(smoke_inf_1[0][1], 3), fmt(smoke_inf_2[0][1], 3), fmt(smoke_inf_3[0][1], 3)],
    '烟幕干扰弹投放点的z坐标 (m)': [fmt(smoke_inf_1[0][2], 3), fmt(smoke_inf_2[0][2], 3), fmt(smoke_inf_3[0][2], 3)],
    '烟幕干扰弹起爆点的x坐标 (m)': [fmt(smoke_inf_1[1][0], 3), fmt(smoke_inf_2[1][0], 3), fmt(smoke_inf_3[1][0], 3)],
    '烟幕干扰弹起爆点的y坐标 (m)': [fmt(smoke_inf_1[1][1], 3), fmt(smoke_inf_2[1][1], 3), fmt(smoke_inf_3[1][1], 3)],
    '烟幕干扰弹起爆点的z坐标 (m)': [fmt(smoke_inf_1[1][2], 3), fmt(smoke_inf_2[1][2], 3), fmt(smoke_inf_3[1][2], 3)],
    '有效干扰时长 (s)': [fmt(t_spans[0], 2), fmt(t_spans[1], 2), fmt(t_spans[2], 2)],
}

# 在保存Excel前加上用户确认
user_input = input("是否保存结果到Excel？输入Y或y保存，其他任意键不保存：")
if user_input.strip().lower() == 'y':
    workbook = load_workbook(excel_file_path)
    sheet = workbook.active

    # 获取列标题到列索引的映射
    column_map = {}
    for col_idx, cell in enumerate(sheet[1], 1):
        column_map[cell.value] = get_column_letter(col_idx)

    # 确定数据填充的起始行（第一行为标题行）
    start_row_index = 2 # Excel 行从 1 开始，标题占一行，数据从第 2 行开始

    # 准备要填充的数据，确保所有列表长度一致
    max_data_rows = 0
    for col_data in data_to_fill.values():
        max_data_rows = max(max_data_rows, len(col_data))

    for col_name in data_to_fill:
        while len(data_to_fill[col_name]) < max_data_rows:
            data_to_fill[col_name].append(None)

    # 填充数据
    for col_header, col_data in data_to_fill.items():
        if col_header in column_map:
            column_letter = column_map[col_header]
            for i, value in enumerate(col_data):
                target_row = start_row_index + i
                cell = sheet[f"{column_letter}{target_row}"]
                cell.value = value
        else:
            print(f"警告：列 '{col_header}' 在 Excel 文件中未找到，数据将不会填充。")
            
    # 保存修改后的 Excel 文件
    workbook.save(excel_file_path)
    print(f"成功使用新数据更新了 '{excel_file_path}'。")
else:
    print("未保存结果到Excel。")